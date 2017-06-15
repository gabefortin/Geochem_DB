import os, sys, csv, pyodbc, ogr, osr
from openpyxl import load_workbook

db_path = 'C:\\Project\\ARIS_Geochem_dev\\data\\ARIS_geochem.stage.accdb'
data_dir = 'C:\\Project\\ARIS_Geochem_dev\\data\\_AR Data Staging Location\\'

print db_path
#Collect all xls file name under the specified directory
xls_list = os.listdir(data_dir)
print xls_list
#Loop through each file in the drectory
for f in range(len(xls_list)):
    xls_name = data_dir + xls_list[f]
    print xls_name
    #Open and exam each xls file
    wb = load_workbook(filename = xls_name)
    ws = wb[wb.sheetnames[0]]
    print xls_name + ' is being loaded ...'