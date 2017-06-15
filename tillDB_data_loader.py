# -*- coding: cp1252 -*-
'''+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
   This script is coded to load and update the following 3 tables: 'data_sample', 'data_analyte', and
   'data_publish' of the till geochem database (TillDB) using the data provided in the staged xlsx
   files.
   
   Input
         1) Path to the directory containing the staged xlsx files
         2) Path to the till geochem database, TillDB (MS Access)
         
  Output
         Data inserted into the above 3 tables in the TillDB
        
  Additional info
         1) Put all the staged xlsx files in a single directory. Data published in each publication
            have to be staged into a single staged xlsx file.
            
         2) In case a set of samples appeared in multiple publications, the corresponding staged xlsx
            file for each publication should be prepared separately with the same set of samples.
            The relationship between these samples and the corresponding publications is specified in
            'data_publish' table. To avaid duplication, these samples are saved only once in the
            'data_sample' table of the database.

         3) It is assumed that the data in these xlsx files are error free. All errors have
            to be addressed during the data screening process. So it is required to run
            'data_screener.py' (before execution of this script) to check whether data are properly
            staged. 
         
         4) Prior to running this script, the 3 code tables in the TillDB, including 'code_unit',
            'code_method', and 'code_lab', need to be updated manually based on the information
            provided in the staged xlsx files (to be loaded).

         6) All entries in the staged xlsx files are assumed to be in plain text. So make sure
            no subscript or superscript exists.
            
         7) This script only works when the input data are organized strictly following the exact
            format defined in the data staging.
            
         8) The current algorithm can not handle the following situations:
            a) Rounded concentration values. For example, Au is an analyte of sample_A with determination
            of 2.1 ppb by FA in lab_A and reported in report_A. A few years later, sample_A was
            published again where the Au value was rounded to 2.0 ppb. In this situation, two Au rows
            will be created in 'data_analyte' table. This causes redundency. 

            b) Same re-analysis values. For example, Au is an analyte of sample_A with determination
            of 2.1 ppb by FA in lab_A and reported in report_A. A few year later, the same sample
            was sent to lab_A (again) for analysis for Au using FA (again), and found the same result
            (2.1 ppb). In this situation, only one Au is recorded in 'data_analyte' table. In this case,
            these 2 Au values should be both recorded (as 2 rows) in 'data_analyte'. 
            
            c) Same samples republished using different 'sample_name'. For example, sample_A was originally
            published in 1988. It was re-published in 1990 as sample_B. In this case, both sample_A
            and sample_B will be recorded as 2 rows in 'data_sample' table and all their analytes included
            will be recorded in 'data_analyte' table. This causes redundency in both tables.

            All these problems are difficult to spot in data screening. They need to be examined visually in
            the final data products.
  Status
      Operational

  Developer
      T. Han
      
  Last update
      2017-03-17
  +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++'''
import os, sys, csv, pyodbc, ogr, osr
from openpyxl import load_workbook
#========================================== Sub-routines =================================================
#Get next row number of a given field and table
#Syntax: get_rownum(db_cursor, string, string) return int
def get_rownum(db_cur, fld_name, tab_name):
    db_cur.execute('select max(%s) from %s' %(fld_name, tab_name))

    max_val = db_cur.fetchone()
    if max_val[0] == None:
        return 1
    else:
        return int(max_val[0]) + 1
    
#Get unit_id with the given unit_name as defined in 'code_unit' table
#Syntax: get_unitid (db_cursor, list) return list
def get_unitid(db_cur, name_list):
    temp_id = []
    temp_name = []
    db_cur.execute("""select unit_id, name from code_unit""")
    all_records = db_cur.fetchall()
    for record in all_records:
        temp_id.append(record[0])
        temp_name.append(record[1])
        
    unitid_list = []
    for i in range(len(name_list)):
        name_list[i] = name_list[i].lower()

        #Standarize unit name
        if name_list[i].lower() == 'g':
            name_list[i] = 'g'

        #Validate
        if name_list[i] in temp_name:
            indx = temp_name.index(name_list[i])
            unitid_list.append(temp_id[indx])
        else:
            print 'Invalid unit: ' + name_list[i] 
            sys.exit()

    return unitid_list
#============================================= Main routine =========================================
def main():   
    #File path
    db_path = 'C:\\Project\\TillDB\\data\\tillDB_curr.accdb'
    data_dir = 'C:\\Project\\TillDB\\data\\workspace\\'

    #Database connection
    db_conn = pyodbc.connect('Driver={Microsoft Access Driver (*.mdb, *.accdb)};DBQ='+db_path)
    cur = db_conn.cursor()

    #Get next id values from tables: 'data_sample', 'data_analyte', and 'data_publish' 
    sample_id = get_rownum(cur, 'sample_id', 'data_sample')
    analyte_id = get_rownum(cur, 'analyte_id', 'data_analyte')
    pub_id = get_rownum(cur, 'pub_id', 'data_publish')
    
    #Collect all xls file name under the specified directory
    xls_list = os.listdir(data_dir)

    #Loop through each file in the drectory
    for f in range(len(xls_list)):
        xls_name = data_dir + xls_list[f]

        #Open and exam each xls file
        wb = load_workbook(filename = xls_name)
        ws = wb[wb.sheetnames[0]]
        print xls_name + ' is being loaded ...'

        #Extract the top 6 rows from 'XXXXX.xlsx'
        analyte_list = list()
        unit_list = list()
        mdl_list = list()
        method_list = list()
        labid_list = list()
        size_list = list()
        for c in range(20, ws.max_column + 1):
            analyte_list.append(str(ws.cell(row = 1, column = c).value).replace(' ', ''))
            unit_list.append(str(ws.cell(row = 2, column = c).value).replace(' ', ''))
            mdl_list.append(str(ws.cell(row = 3, column = c).value).replace(' ', ''))
            method_list.append(str(ws.cell(row = 4, column = c).value).replace(' ', ''))
            labid_list.append(str(ws.cell(row = 5, column = c).value).replace(' ', ''))
            size_list.append(str(ws.cell(row = 6, column = c).value).replace(' ', ''))

        #Get 'unit_id' for the retrieved unit names
        unitid_list = get_unitid(cur, unit_list)

        #Step through the remaining rows (values are all read in as strings)
        for r in range(8, ws.max_row + 1):
            #Retrieve 'pub_issue'
            pub_issue = str(ws.cell(row = r, column = 18).value).replace(' ', '')
            
            #Retrieve analyte values
            abundance = list()
            for c in range(20, ws.max_column + 1):
                abundance.append(str(ws.cell(row = r, column = c).value).replace(' ', ''))
            #-------------------------------------- Update the 3 relevent tables  -----------------------------------
            sample_code = str(ws.cell(row = r, column = 2).value).replace(' ', '')
            cur.execute("""select sample_id from data_sample where sample_code = ?""", sample_code)
            check_sample = cur.fetchone()
            
            #The current sample is already in 'data_sample' table (which won't be updated).
            if check_sample <> None:
                #----------------------------- Update 'data_publish' table if applicable ----------------------------
                cur.execute("""select count(*) from data_publish where (sample_id = ?) and (pub_issue = ?)""",
                               check_sample[0], pub_issue)
                rec_count = cur.fetchone()
                if rec_count[0] == 0:
                    cur.execute("""insert into data_publish values (?, ?, ?)""", pub_id, pub_issue, check_sample[0])
                    cur.commit()
                    pub_id = pub_id + 1
                #----------------------------- Update 'data_analyte' table if applicable ----------------------------    
                for i in range(len(analyte_list)):
                    cur.execute("""select count(*) from data_analyte where (analyte = ?) and (abundance = ?) and
                               (size_frac = ?) and (unit_id = ?) and (method_id = ?) and  (lab_id = ?) and
                               (sample_id = ?)""", analyte_list[i], abundance[i], size_list[i], int(unitid_list[i]), \
                                int(method_list[i]), int(labid_list[i]), check_sample[0])
                    check_analyte = cur.fetchone()
                    
                    #Update 'data_analye' table if applicable
                    if (check_analyte[0] == 0) and (abundance[i] <> '') and (abundance[i] <> 'None'): 
                        cur.execute("""insert into data_analyte values (?, ?, ?, ?, ?, ?, ?, ?, ?)""", analyte_id, \
                                    analyte_list[i], abundance[i], mdl_list[i], size_list[i], int(unitid_list[i]), \
                                    int(method_list[i]), int(labid_list[i]), check_sample[0])
                        cur.commit()
                        analyte_id = analyte_id + 1
                        
            #The current sample is not in 'data_sample' table (which is to be updated)
            else:
                sample_name = str(ws.cell(row = r, column = 1).value).replace(' ', '')
                if sample_name == '':
                    sample_name = 'NA'
                    
                sample_type = str(ws.cell(row = r, column = 3).value).replace(' ', '')
                if sample_type == '':
                    sample_type = 'NA'

                depth = str(ws.cell(row = r, column = 4).value).replace(' ', '')
                if depth == '':
                    depth = 'NA'

                duplicate = str(ws.cell(row = r, column = 5).value).replace(' ', '')
                if duplicate == '':
                    duplicate = 'NA'

                borehole = str(ws.cell(row = r, column = 6).value).replace(' ', '')
                if borehole == '':
                    borehole = 'NA'
                
                core_top = str(ws.cell(row = r, column = 7).value).replace(' ', '')
                if core_top == '':
                    core_top = 'NA'

                core_bottom = str(ws.cell(row = r, column = 8).value).replace(' ', '')
                if core_bottom == '':
                    core_bottom = 'NA'

                azimuth = str(ws.cell(row = r, column = 9).value).replace(' ', '')
                if azimuth == '':
                    azimuth = 'NA'

                dip = str(ws.cell(row = r, column = 10).value).replace(' ', '')
                if dip == '':
                    dip = 'NA'

                drill_type = str(ws.cell(row = r, column = 11).value).replace(' ', '')
                if drill_type == '':
                    drill_type = 'NA'

                material_type = str(ws.cell(row = r, column = 12).value)
                if material_type == '':
                    material_type = 'NA'

                sample_desc = str(ws.cell(row = r, column = 13).value)
                if sample_desc == '':
                    sample_desc = 'NA'
                    
                x_coord = str(ws.cell(row = r, column = 14).value).replace(' ', '')
                y_coord = str(ws.cell(row = r, column = 15).value).replace(' ', '')
                
                z_coord = str(ws.cell(row = r, column = 16).value).replace(' ', '')
                if z_coord == '':
                    z_coord = 'NA'
                    
                epsg_srid = str(ws.cell(row = r, column = 17).value).replace(' ', '')
                coord_conf = str(ws.cell(row = r, column = 19).value).replace(' ', '').upper()
                
                #Assemble a row of values to be written to 'data_sample' table
                sample_values = [sample_id, sample_code, sample_name, sample_type, depth, duplicate, borehole,
                                 core_top, core_bottom, azimuth, dip, drill_type, material_type,
                                 sample_desc, x_coord, y_coord, z_coord, int(epsg_srid), coord_conf]

                #----------------------------- Add to the "data_sample" table ------------------------------------------
                cur.execute("""insert into data_sample values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                               sample_values)
                cur.commit()
                #----------------------------- Add to the "data_analyte" table -----------------------------------------
                for a in range(len(analyte_list)):
                    if (abundance[a] <> '') and (abundance[a] <> 'None'):
                        analyte_values = [analyte_id, analyte_list[a], abundance[a], mdl_list[a], size_list[a],
                                          int(unitid_list[a]), int(method_list[a]), int(labid_list[a]), sample_id]
                        cur.execute("""insert into data_analyte values (?, ?, ?, ?, ?, ?, ?, ?, ?)""", analyte_values)
                        cur.commit()
                        analyte_id = analyte_id + 1
                #---------------------------------- Update 'data_publish' table ----------------------------------------
                cur.execute("""select count(*) from data_publish where (sample_id = ?) and (pub_issue = ?)""",
                           sample_id, pub_issue)
                rec_count = cur.fetchone()
                if rec_count[0] == 0:
                    cur.execute("""insert into data_publish values (?, ?, ?)""", pub_id, pub_issue, sample_id)
                    cur.commit()
                    pub_id = pub_id + 1

                sample_id = sample_id + 1
            
    db_conn.close() 
    print 'Job done!'
    
if __name__ == "__main__":
    main()
