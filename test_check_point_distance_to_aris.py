# -*- coding: cp1252 -*-
'''++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
   This script is modified from Tian Han's data_screener.py script. It has been adapted to
   check sample location information for the ARIS Geochem data compilation project.
   
   This script is coded to identify and report problems found in the staged geochem sample
   location data in xlsx files before loading them loaded into the ARIS Geochem database.
   Please beware that this script only identifies and reports problems. The problems identified
   need to be fixed manually.
   
   The file names for the xlsx sheets are used to code the ARIS numbers into the database, so
   the names are important. All files should be name XXXXX_loc.xlsx where XXXXX is the ARIS
   report number.
   
   Input
         1) Path to the directory where the staged xlsx files reside
         2) aris_geochem Database path (this is not currently necessary as of June 1, 2017
         but may be used at a later date).
         3) Some hardcoded parameters 
         
   Output
         Display identified problems

   Operation note
         1) The script has many independent code-blocks targeting different types of errors as listed
            below. Though each code-block can be executed randomly, it is strongly recommended to
            run them sequentially in the given order, one at a time. For example, it should start with
            checking format related problems in the xlsx files by commenting out all other code-blocks
            which are for checking other types of errors. Once the format-related errors are identified
            and fixed, the format-checking code-block should be commented out and then move on to the
            "analyte name checking block" ...

         2) Duplicate samples should be given different "sample_code". Re-published samples should
            have the same 'sample_code' across all publications. Sections 15 and 16 are designed to
            ensure if the above requirements are met.
            
    This script is able to check the following:
         1)  Check the format of the spreadsheet to ensure that all necessary columns are present, properly named and in
            the right order.
         2) Check that x, y, z coordinates and epsg codes are present and are in fact numeric. Check that z values, when
            present are between 0 and 3000 metres.
         3) Check that values entered in the sample_type column match the pre-determined options. The script will 
            automatically turn any upper case character to lower case.
         4)  Check that values in sample_subtype match accepted values. This column is case sensitive.
         5)  Check that values entered in coord_conf match accepted values.
         6)  Confirm that points actually plot within the province of BC by computing NAD83 Lat Long on the fly and
            comparing to a shapefile with the BC boundary.
        7)  Check that all dates under sample_date are proper dates.
         
    
  Status
      In Development

  Future improvements
        The algorithms used in 16) and 17) above generate too many false alarms. They are not helpful
        enough to identify duplicate samples.

  Developers
      T. Han, G. Fortin
      
  Last update
      2017-05-21
  +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++'''
import os, ogr, osr, datetime, openpyxl
from openpyxl import load_workbook
from dateutil.parser import parse

# This function is to check if a given string can be converted to a decimal number
# Syntax: is_number(string) return logic
def is_number(n):
    try:
        float(n)
        return True
    except ValueError:
        return False
    
# This function is to check if a given variable can be converted to a string
# Syntax: is_number(variable) return logic
def is_string(s):
    try:
        str(s)
        return True
    except ValueError:
        return False
    
#Project input coordinates to NAD83 lat. and long (EPSG_SRID = 4269)
#Syntax: project2nad83(float, float, int) returns [nad83_long, nad83_lat]
def project2nad83 (x_coord, y_coord, source_epsg):

    #Create a point geometry using the given coordinates
    point = ogr.Geometry(ogr.wkbPoint)
    point.AddPoint(x_coord, y_coord)

    #Specify source spatial reference and projection
    source_ref = osr.SpatialReference()
    source_ref.ImportFromEPSG(source_epsg)

    #Define target spatial reference and projection based on input
    target_ref = osr.SpatialReference()
    target_ref.ImportFromEPSG(4269)
        
    #Perform transformation
    transform = osr.CoordinateTransformation(source_ref, target_ref)
    point.Transform(transform)
    
    return [str(point.GetX(0)), str(point.GetY(0))]
    
def main():   
    #File path
    data_dir = 'C:\\Project\\ARIS_Geochem_dev\\data_testing\\_AR Data Staging Location\\'
    chkrpt_nm = 'C:\\Project\\ARIS_Geochem_dev\\data_testing\\checkreports\\SampleInfoCheckReport_' + \
        datetime.datetime.now().strftime("%Y_%m_%d_%H_%M") + '.xlsx'
    min_aris_path = 'C:\\min_aris\\min_aris.mdb'

    #Database connection
    db_conn = pyodbc.connect('Driver={Microsoft Access Driver (*.mdb, *.accdb)};DBQ='+min_aris_path)
    cur = db_conn.cursor()

    #Minimum long. and lat. difference between sample locations in degree
    min_lon = 0.001
    min_lat = 0.001
    
    #Collect year sub-directories under given "data_dir" 
    xls_list = os.listdir(data_dir)

    #List of allowed EPSG codes
    epsg_list = ['2955','3156','3157','4269','4326','26709','26710']

    #List of allowed sample_types
    samptype_list = ['soil','soil-mmi','silt','stream sediment','till','moss mat']

    #List of allowed sample_subtypes
    subtype_list = ['A Horizon','B Horizon','C Horizon','Ah Horizon','A-B Horizons','B-C Horizons','A-C Horizons',
                    'silt','pit','trench','pan concentrate']

    #create check report xlsx file and 'open' it
    chkwb = openpyxl.Workbook()
    chkwb.remove_sheet(chkwb.get_sheet_by_name('Sheet'))
    chkws = chkwb.create_sheet(title='CheckResults')
    chkws.cell(row=1, column=1).value = 'File Name'
    chkws.cell(row=1, column=2).value = 'Check Type'
    chkws.cell(row=1, column=3).value = 'Problem'
    chkws.cell(row=1, column=4).value = 'Row'
    chkws.cell(row=1, column=5).value = 'Column'

    #function to write problems to check report xlsx
    def write_rpt_err(file,check,problem,row,column):
        chkws.cell(row=chkws.max_row + 1, column=1).value = file
        chkws.cell(row=chkws.max_row, column=2).value = check
        chkws.cell(row=chkws.max_row, column=3).value = problem
        chkws.cell(row=chkws.max_row, column=4).value = row
        chkws.cell(row=chkws.max_row, column=5).value = column



    #================== 8. Check points are within 10km of ARIS point=============
    print '\n8. Check points are within 10km of ARIS point  ...'
    #Looking for samples that do not fall in BC to identify possible coordinate issues

    #Loop through the xls files
    xls_sample = list()  #Build a list for sample_code in xlsx files
    xls_xc = list()      #Build a list for x_coord in xlsx files
    xls_yc = list()      #Build a list for y_coord in xlsx files
    xls_file = list()    #Build a list for xlsx file names
    xls_row = list()
    xls_loccheck = list()
    for f in range(len(xls_list)):
        xls_name = data_dir + xls_list[f]

        #Open and exam each xls file
        wb = load_workbook(filename = xls_name)
        ws = wb[wb.sheetnames[0]]

        if (ws.cell(row = ws.max_row, column = 1)).value is None:
            reallastrow = ws.max_row - 1
        else:
            reallastrow = ws.max_row


        for r in range(2, reallastrow + 1):
            cell_sample = str(ws.cell(row = r, column = 1).value).replace(' ', '')
            cell_xc = str(ws.cell(row = r, column = 9).value).replace(' ', '')
            cell_yc = str(ws.cell(row = r, column = 10).value).replace(' ', '')
            cell_epsg = str(ws.cell(row = r, column = 12).value).replace(' ', '')
            cell_row = str(r)

            xls_sample.append(cell_sample)

            if cell_epsg <> '4269': #NAD83 geographic
                [cell_xc, cell_yc] = project2nad83 (float(cell_xc), float(cell_yc), int(cell_epsg))
            xls_xc.append(cell_xc)
            xls_yc.append(cell_yc)
            xls_file.append(xls_name)
            xls_row.append(cell_row)


    #Loop through to check each sample coord
    for i in range(len(xls_sample)):
        #Examine sample locations

        # create point geometry
        pt = ogr.Geometry(ogr.wkbPoint)
        pt.SetPoint_2D(0, float(xls_xc[i]), float(xls_yc[i]))

        aris_pt = ogr.Geometry(ogr.wkbPoint)
        aris_pt.SetPoint_2D(0,-121.793333,52.520556)
        inSpatialRef = osr.SpatialReference()
        inSpatialRef.ImportFromEPSG(4269)
        outSpatialRef = osr.SpatialReference()
        outSpatialRef.ImportFromEPSG(3153)
        coordTransform = osr.CoordinateTransformation(inSpatialRef,outSpatialRef)
        pt.Transform(coordTransform)
        aris_pt.Transform(coordTransform)

        print pt.Distance(aris_pt)

    chkwb.save(chkrpt_nm)  # save results to the xlsx report

    print '\nJob done.'
    
if __name__ == "__main__":
    main()
