# -*- coding: cp1252 -*-
'''+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
   This script is modified from Tian Han's tillDB_data_screener script. It is coded to identify and report problems
   found in the staged ARIS geochem data results xlsx files   before loading them loaded into the tillDB database.
   Please beware that this script only identifies and reports problems. The problems identified need to be fixed
   manually.
   
   Input
         1) Path to the directory where the staged results xlsx files reside
         2) ARIS Geochem stage Database path
         3) Some hardcoded parameters
         
   Output
         Display identified problems

   Operation note
         1) The script has many independent code-blocks targeting different types of errors as listed
            below. Though each code-block can be executed randomly, it is strongly recommended to
            run them sequentially in the given order, one at a time. For example, it should start with
            checking format related problems in the xlsx files by commenting out all other code-blocks
            which are for checking other types of errors. Once the format-related errors are identified
            and fixed, the format-checking code-block should be commented out and then move on to the
            "analyte name checking block" ...

         2) Duplicate samples should be given different "sample_code". Re-published samples should
            have the same 'sample_code' across all publications. Sections 15 and 16 are designed to
            ensure if the above requirements are met.
            
    This script is able to check the following:
         1)  if format of staged xlsx files is correct;
         2)  if analyte name are appropriate;
         3)  if analyte value unit exists in database;
         4)  if detection limit value present in xlsx files;
             Be noted that blank detection limit is ok. This entry is optional.
             
         5)  if analytical "method_id" exists in database;
         6)  if "lab_id" exists in database;
         7)  if "size_fraction" is blank, which is not allowed;
         8)  if duplicate columns exist within xlsx files; 
             It should be noted that duplicate columns may exist. For example, an analyte was initally
             and re-analyzed using the same method by the same lab. It often happens to Au which results
             in 2 Au columns. Extra caution should be exersized when dealing with these cases. 

         9)  if duplicate or empty samples exist within each xlsx file;
         10) if invalid x_coord, y_coord, z_coord, and EPSG_SRID values exist in xlsx files;
         11) if xlsx file name corresponds to "pub_issue" in xlsx files; 
         12) if "coord_conf" values are appropriate in xlsx files;
         13) if analyte value units are appropriate for some method-dependent analytes;
         14) if method-specific analytes use the right methods;
         15) if certain analyte values are appropriate

         16) if duplicate samples exist among the samples in the xlsx files.
             Duplicate-sample flag will be triggered if the following 2 conditions are met simultaneously:
             a) closely located;
             b) similar sample names.

             Be noted that 1) the flagged samples may be those that are re-published; 2) this algorithm
             doesn't work if 2 samples are named very differently through they are true duplicates; and
             3) the flagged samples are not necessarily true duplicates. They need to examined carefully.
             
         17) if duplciate samples exist among the samples in the xlsx files and those in the database.
             Duplicate-sample flag will be triggered if the following 2 conditions are met simultaneously:
             a) closely located;
             b) similar sample names.

             Be noted that 1) the flagged samples may be those that are re-published; 2)this algorithm
             doesn't work if 2 samples are named very differently through they are true duplicates; and
             3) the flagged samples are not necessarily true duplicates. They need to examined carefully.
    
  Status
      Operational

  Furture improment
        The algorithms used in 16) and 17) above generate too many false alarms. They are not helpful
        enough to identify duplicate samples.

  Developer
      T. Han
      
  Last update
      2016-12-15
  +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++'''
import os, sys, csv, pyodbc, ogr, osr
from openpyxl import load_workbook

# This function is to check if a given string can be converted to a decimal number
# Syntax: is_number(string) return logic
def is_number(n):
    try:
        float(n)
        return True
    except ValueError:
        return False
    
# This function is to check if a given variable can be converted to a string
# Syntax: is_number(variable) return logic
def is_string(s):
    try:
        str(s)
        return True
    except ValueError:
        return False
    
#Project input coordinates to NAD83 lat. and long (EPSG_SRID = 4269)
#Syntax: project2nad83(float, float, int) returns [nad83_long, nad83_lat]
def project2nad83 (x_coord, y_coord, source_epsg):

    #Create a point geometry using the given coordinates
    point = ogr.Geometry(ogr.wkbPoint)
    point.AddPoint(x_coord, y_coord)

    #Specify source spatial reference and projection
    source_ref = osr.SpatialReference()
    source_ref.ImportFromEPSG(source_epsg)

    #Define target spatial reference and projection based on input
    target_ref = osr.SpatialReference()
    target_ref.ImportFromEPSG(4269)
        
    #Perform transformation
    transform = osr.CoordinateTransformation(source_ref, target_ref)
    point.Transform(transform)
    
    return [str(point.GetX(0)), str(point.GetY(0))]
    
def main():   
    #File path
    db_path = 'C:\\Project\\ARIS_Geochem_dev\\data\\ARIS_geochem_stage.accdb'
    data_dir = 'C:\\Project\\ARIS_Geochem_dev\\data\\_AR Data Staging Results\\'

    #Database connection
    db_conn = pyodbc.connect('Driver={Microsoft Access Driver (*.mdb, *.accdb)};DBQ='+db_path)
    cur = db_conn.cursor()

    #Standard list of analyte elements (provided by Alexei).
    element_list = ['Si','Al','Ca','Fe','K','Mg','Na','P','S','SO3','CO2','Sn','Sr','SiO2',\
                    'TiO2', 'Al2O3','Fe2O3(T)','FeO(T)','FeO','MnO','MgO','CaO','Na2O','K2O','P2O5',\
                    'V2O5','Cr2O3','CoO','NiO','CuO','ZnO','As2O3','SrO','Y2O3','ZrO2','Nb2O5','MoO2',\
                    'SnO','Sb2O3','BaO','TR2O3','La2O3','Lu2O3','Ta2O5','Ti','WO2','PbO','Bi2O3','H2O+',\
                    'H2O-','LOI','Total','Cl','Au','Pd','Pt','Re','Os','Ir','Ru','Rh','Hg','Ag', \
                    'Wt','As','B','Ba','Be','Bi','Br','Cd','Ce','Co','Cr','Cs','Cu','Dy','Er','Eu',\
                    'Ga','Gd','Ge','Hf','Ho','In','La','Li','Lu','Mo','Mn','Nb','Nd','Ni','Pb','Pr',\
                    'Rb','Sb','Sc','Se','Sm','Ta','Tb','Te','Th','Tl','Tm','U','V','W','Y','Yb',\
                    'Zn','Zr', 'F', 'C(T)', 'C(org)', 'C(inorg)', 'H2O(T)', 'N', 'CaF2', 'Wgt']

    #Standard list of unit corresponding to the analyte above (provided by Alexei)
    standard_unit = ['%', '%', '%', '%', '%', '%', '%', '%', '%', '%', '%', 'ppm', 'ppm', '%', \
                     '%', '%', '%', '%', '%', '%', '%', '%', '%', '%', '%', \
                     '%', '%', '%', '%', '%', '%', '%', '%', '%', '%', '%', '%', \
                     '%', '%', '%', '%', '%', '%', '%', '%', '%', '%', '%', '%', \
                     '%', '%', '%', '%', 'ppb', 'ppb', 'ppb', 'ppb', 'ppb', 'ppb', 'ppb', 'ppb', 'ppb', 'ppm', \
                     'g', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', \
                     'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', \
                     'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', \
                     'ppm', 'ppm', 'ppm', '%', '%', '%', '%', 'ppm', '%', 'kg']

    #List of analytes who units are method-dependent
    dependent_analyte = ['Ag', 'Au', 'Hg', 'Sn', 'Sr']
    dependent_method =  ['AIP', 'MIP', 'INA', 'INA', 'INA']
    dependent_unit =    ['ppm', 'ppb', 'ppb', 'ppm', 'ppm']
    
    #Standard_list for analytes without detection limit
    nolimit_list = ['Total', 'Wt']

    #List of fixed combo between analyte and method
    fix_analyte = ['LOI', 'H2O-', 'H2O(T)', 'C(T)', 'F', 'FeO', 'CO2', 'C(org)']
    fix_method =  ['7', '7', '11', '11', '20', '21', '11', '11']

    #Collect year sub-directories under given "data_dir"
    xls_list = os.listdir(data_dir)
    #====================================== 1. Check xls file format ======================================

    print '1. Examine file format ...'
    #Loop through xls file
    for f in range(len(xls_list)):
        xls_name = data_dir + xls_list[f]
            
        #Open and exam each xls file
        wb = load_workbook(filename = xls_name)
        ws = wb[wb.sheetnames[0]]
        if (str(ws.cell(row = 5, column = 1).value)).lower() <> 'sample_name':
            print '    ' + xls_list[f] + ': Wrong Sample_Name column header'
        if (str(ws.cell(row = 5, column = 2).value)).lower() <> 'cert_no':
            print '    ' + xls_list[f] + ': Wrong Cert_No column header'

        if (str(ws.cell(row = 1, column = 2).value)).lower() <> 'analyte':
            print '    ' + xls_list[f] + ': Wrong Analyte row header'
        if (str(ws.cell(row = 2, column = 2).value)).lower() <> 'unit':
            print '    ' + xls_list[f] + ': Wrong Unit row header'
        if (str(ws.cell(row = 3, column = 2).value)).lower() <> 'd_limit':
            print '    ' + xls_list[f] + ': Wrong D_Limit row header'
        if (str(ws.cell(row = 4, column = 2).value)).lower() <> 'method_id':
            print '    ' + xls_list[f] + ': Wrong Method_ID row header'

    #===================================== 2. Check analyte name =======================================
    print '\n2. Examine analyte names ...'
    #Loop through xls file
    for f in range(len(xls_list)):
        xls_name = data_dir + xls_list[f]
            
        #Open and exam each xls file
        wb = load_workbook(filename = xls_name)
        ws = wb[wb.sheetnames[0]]

        reallastcolumn = ws.max_column
        for i in range(ws.max_column,3,-1):
            if (ws.cell(row=1, column=i)).value is None:
                reallastcolumn = i - 1

        for c in range(3, reallastcolumn + 1):
            if str(ws.cell(row = 1, column = c).value) not in element_list:
                print '    ' + xls_list[f] + ', Column ' + str(c) + ': ' + str(ws.cell(row = 1, column = c).value) + \
                      ' analyte name not in list, check name'

    #=============================== 3. Verify analyte unit against database ===========================

    #should check if units are lower case, if not, convert to lower case before running checks. Alternatively
    # use .lower to run check then convert to lower case in import script.

    print '\n3. Examine unit name ...'
    unit_list = list()  #Build a list of unit name using current database values
    cur.execute("""select name from code_unit""")
    val_rows = cur.fetchall()
    for i in range(len(val_rows)):
        unit_list.append(str(val_rows[i][0]))
    unit_list.pop(0)    #Remove 1st item: 'unknown'

    #Loop through xls file
    for f in range(len(xls_list)):
        xls_name = data_dir + xls_list[f]
            
        #Open and exam each xls file
        wb = load_workbook(filename = xls_name)
        ws = wb[wb.sheetnames[0]]

        reallastcolumn = ws.max_column
        for i in range(ws.max_column, 3, -1):
            if (ws.cell(row=1, column=i)).value is None:
                reallastcolumn = i - 1

        for c in range(3, reallastcolumn + 1):
            if str(ws.cell(row = 2, column = c).value) not in unit_list:
                print '    ' + xls_list[f] + ', Column ' + str(c) + ': ' + str(ws.cell(row = 2, column = c).value) + \
                      ' unit not in list, check units'

    #==================================== 4. Check detection limit ======================================
    #Detection limit is not mandatory, which can be left blank.
    #Loop through xls files
    print '\n4. Examine detection limit ...'
    for f in range(len(xls_list)):
        xls_name = data_dir + xls_list[f]
            
        #Open and exam each xls file
        wb = load_workbook(filename = xls_name)
        ws = wb[wb.sheetnames[0]]

        reallastcolumn = ws.max_column
        for i in range(ws.max_column, 3, -1):
            if (ws.cell(row=1, column=i)).value is None:
                reallastcolumn = i - 1

        for c in range(3, reallastcolumn + 1):
            dlimit_cell = str(ws.cell(row = 3, column = c).value)
            analyte_cell = str(ws.cell(row = 1, column = c).value)
            if (dlimit_cell == 'None') or dlimit_cell.isspace():
                if analyte_cell not in nolimit_list:
                    print '    ' + xls_list[f] + ', Column ' + str(c) + ': ' + dlimit_cell + ' missing detection limit'
    #==================================== 5. Check method_id ======================================
    print '\n5. Examine method_id ...'
    method_list = list()   #Build a list of method_id using current database values
    cur.execute("""select method_abbr from code_method""")
    val_rows = cur.fetchall()
    for i in range(len(val_rows)):
        method_list.append(str(val_rows[i][0]))
    method_list.pop(0)     #Remove 1st item: 'unknown'
       
    #Loop through xls files
    for f in range(len(xls_list)):
        xls_name = data_dir + xls_list[f]
            
        #Open and exam each xls file
        wb = load_workbook(filename = xls_name)
        ws = wb[wb.sheetnames[0]]

        reallastcolumn = ws.max_column
        for i in range(ws.max_column, 3, -1):
            if (ws.cell(row=1, column=i)).value is None:
                reallastcolumn = i - 1

        for c in range(3, reallastcolumn + 1):
            if str(ws.cell(row = 4, column = c).value) not in method_list:
                print '    ' + xls_list[f] + ', Column ' + str(c) + ': ' + str(ws.cell(row = 4, column = c).value) + \
                      ' method not in DB, check method'

    #================= 6. Check duplicate columns (i.e. same analyte, method, and lab) =============
    # It should be noted that duplicate columns may exist. For example, an analyte was initally
    # and re-analyzed using the same method by the same lab and with the same size fration. This will
    # result in 2 columns for this analyte. Extra caution should be excersized when dealing with these
    # cases.
    print '\n6. Examine duplicate columns within each xls file ...'
    #Loop through xls files
    for f in range(len(xls_list)):
        xls_name = data_dir + xls_list[f]
            
        #Open and exam each xls file
        wb = load_workbook(filename = xls_name)
        ws = wb[wb.sheetnames[0]]

        work_analyte = []
        work_method = []

        reallastcolumn = ws.max_column
        for i in range(ws.max_column, 3, -1):
            if (ws.cell(row=1, column=i)).value is None:
                reallastcolumn = i - 1

        for c in range(3, reallastcolumn + 1):
            work_analyte.append(str(ws.cell(row = 1, column = c).value))
            work_method.append(str(ws.cell(row = 4, column = c).value))

        temp_analyte = list(work_analyte)
        for i in range(len(work_analyte)):
            temp_analyte[i] = str(i)
            if work_analyte[i] in temp_analyte:
                indx = temp_analyte.index(work_analyte[i])
                if (work_method[i] == work_method[indx]) :
                    print '    ' + xls_list[f] + ', Column ' + str(c) + ': ' + work_analyte[i] + ' duplicate analyte'

    #===================================== 7. Check method-dependent unit =================================
    print '\n7.Examine method-dependent analyte unit ... (this check is currently ignored)'
    #This should be refined to account for each method's 'allowed' or possible units rather than only one unit
    #type allowed per element.
    '''
    method_list = list()    #Build a list of method_id using current database values
    cur.execute("""select method_abbr from code_method""")
    val_rows = cur.fetchall()
    for i in range(len(val_rows)):
        method_list.append(str(val_rows[i][0]))

    #group_list = list()     #Build a list of method_group using current database values
    #cur.execute("""select method_group from code_method""")
    #val_rows = cur.fetchall()
    #for i in range(len(val_rows)):
    #    group_list.append(str(val_rows[i][0]))

    #Loop through xls file under year directory
    for f in range(len(xls_list)):
        xls_name = data_dir + xls_list[f]
            
        #Open and exam each xls file
        wb = load_workbook(filename = xls_name)
        ws = wb[wb.sheetnames[0]]

        reallastcolumn = ws.max_column
        for i in range(ws.max_column, 3, -1):
            if (ws.cell(row=1, column=i)).value is None:
                reallastcolumn = i - 1

        for c in range(3, reallastcolumn + 1):
            analyte_cell = str(ws.cell(row = 1, column = c).value).replace(' ', '')
            unit_cell = str(ws.cell(row = 2, column = c).value).replace(' ', '')
            method_cell = str(ws.cell(row = 4, column = c).value).replace(' ', '')

            right_unit = standard_unit[element_list.index(analyte_cell)]
            #check_code = group_list[method_list.index(method_cell)]
            #if analyte_cell in dependent_analyte:
            #    if check_code == dependent_method[dependent_analyte.index(analyte_cell)]:
            #        right_unit = dependent_unit[dependent_analyte.index(analyte_cell)]
            
            if unit_cell <> right_unit:
                print '    ' + xls_list[f] +  ', Column ' + str(c) + ': ' + analyte_cell + ' ' + unit_cell + \
                      ' should be ' + right_unit
    '''
    #======================================== 8. Check fixed analyte-method combo ====================================
    #this may not be useful for the aris geochem db... to be revisited
    print '\n8.Examine fixed analyte-method combo ...'
    #Loop through xls file under year directory
    for f in range(len(xls_list)):
        xls_name = data_dir + xls_list[f]
            
        #Open and exam each xls file
        wb = load_workbook(filename = xls_name)
        ws = wb[wb.sheetnames[0]]

        reallastcolumn = ws.max_column
        for i in range(ws.max_column, 3, -1):
            if (ws.cell(row=1, column=i)).value is None:
                reallastcolumn = i - 1

        for c in range(3, reallastcolumn + 1):
            analyte_cell = str(ws.cell(row = 1, column = c).value).replace(' ', '')
            method_cell = str(ws.cell(row = 4, column = c).value).replace(' ', '')
            if analyte_cell in fix_analyte:
                right_method = fix_method[fix_analyte.index(analyte_cell)]
                if method_cell <> right_method:
                    print '    ' + xls_list[f] + ', Column ' + str(c) + ': ' + analyte_cell + ' ' + method_cell + \
                          ' should be ' + right_method

    #========================================== 9. Check analyte value ==============================================
    print '\n9.Examine analyte values ...'
    #Loop through xls file under year directory
    for f in range(len(xls_list)):
        xls_name = data_dir + xls_list[f]
            
        #Open and exam each xls file
        wb = load_workbook(filename = xls_name)
        ws = wb[wb.sheetnames[0]]
            
        freallastcolumn = ws.max_column
        for i in range(ws.max_column,3,-1):
            if (ws.cell(row=1, column=i)).value is None:
                reallastcolumn = i - 1

        for c in range(3, reallastcolumn + 1):
            analyte_cell = str(ws.cell(row = 1, column = c).value).replace(' ', '')
            unit_cell = str(ws.cell(row = 2, column = c).value).replace(' ', '')
            d_limit_cell = str(ws.cell(row = 3, column = c).value).replace(' ', '')

            if (ws.cell(row=ws.max_row, column=1)).value is None:
                reallastrow = ws.max_row - 1
            else:
                reallastrow = ws.max_row

            for r in range(6, reallastrow + 1):
                analyte_value = str(ws.cell(row = r, column = c).value)
                
                if analyte_value == 'None':
                    continue
                elif analyte_value == '0':
                    print '    ' + xls_list[f] + '->' + analyte_cell + ': ' + \
                            analyte_value + '(' + str(r) + ',' + str(c) + ') wrong analyte value'
                elif (not is_number(analyte_value)):
                    if (analyte_value[0] <> '>') and (analyte_value[0] <> '<'):
                        print '    ' + xls_list[f] + '->' + analyte_cell + ': ' + \
                        analyte_value + '(' + str(r) + ',' + str(c) + ') wrong analyte value'
                    elif (analyte_value[0] == '<') and analyte_value[1:] <> d_limit_cell:
                        print '    ' + xls_list[f] + '->' + analyte_cell + ': ' + \
                        analyte_value + '(' + str(r) + ',' + str(c) + ') less than entry does not match detection limit'
                elif (unit_cell == '%') and (is_number(analyte_value)):
                    #Greater-than-100% value is not allowed if anayte unit is % 
                    if (analyte_cell <> 'Total') and float(analyte_value) > 100.0:
                        print '    ' + xls_list[f] + '->' + analyte_cell + ': ' + analyte_value + ' > 100%'

        #Examine rows without any analytic values
        if (ws.cell(row=ws.max_row, column=1)).value is None:
            reallastrow = ws.max_row - 1
        else:
            reallastrow = ws.max_row

        for r in range(6, reallastrow + 1):
            val_count = 0
            reallastcolumn = ws.max_column
            for i in range(ws.max_column, 3, -1):
                if (ws.cell(row=1, column=i)).value is None:
                    reallastcolumn = i - 1

            for c in range(3, reallastcolumn + 1):
                if is_number(str(ws.cell(row = r, column = c).value).replace(' ', '')):
                    val_count = val_count + 1
                elif (analyte_value[0] <> '<') or (analyte_value[0] <> '>'):
                    val_count = val_count + 1
            if val_count == 0:
                print '    ' + xls_list[f] + ': Row = ' + str(r) + ' Blank row without any analytic values'

    # ==================================== 10. Check sample in db ======================================
    print '\n10. Examine sample existence in db for each result ...'
    sample_list = list()  # Build a list of method_id using current database values
    cur.execute("""select samp_dbl_key from vw_sample_dblkey""")
    val_rows = cur.fetchall()
    for i in range(len(val_rows)):
        sample_list.append(str(val_rows[i][0]))

    # Loop through xls files
    for f in range(len(xls_list)):
        xls_name = data_dir + xls_list[f]

        # extract the publication id (i.e. ARIS report number) from the file name
        ar_number = xls_list[f].partition('_')[0]

        # Open and exam each xls file
        wb = load_workbook(filename=xls_name)
        ws = wb[wb.sheetnames[0]]

        if (ws.cell(row=ws.max_row, column=1)).value is None:
            reallastrow = ws.max_row - 1
        else:
            reallastrow = ws.max_row

        for r in range(6, reallastrow + 1):
            sample_name = str(ws.cell(row=r, column=1).value)
            smpdblkey = ar_number + "_" + sample_name

            if smpdblkey not in sample_list:
                print '    ' + xls_list[f] + ': ' + smpdblkey + ' not in DB'

    # ==================================== 11. Check cert_no ======================================
    print '\n11. Examine cert_no to match with filename ...'
    cert_list = list()  # Build a list of method_id using current database values
    cur.execute("""select cert_no from data_cert""")
    val_rows = cur.fetchall()
    for i in range(len(val_rows)):
        cert_list.append(str(val_rows[i][0]))

    # Loop through xls files
    for f in range(len(xls_list)):
        xls_name = data_dir + xls_list[f]

        # extract the cert_no from the file name and compare to that in the sheet
        cert_no_file = xls_list[f].partition('_')[2]
        cert_no_file = cert_no_file[:-13]

        # Open and exam each xls file
        wb = load_workbook(filename=xls_name)
        ws = wb[wb.sheetnames[0]]

        if (ws.cell(row=ws.max_row, column=1)).value is None:
            reallastrow = ws.max_row - 1
        else:
            reallastrow = ws.max_row

        for r in range(6, reallastrow + 1):
            cert_no_sheet = str(ws.cell(row=r, column=2).value)

            if cert_no_file <> cert_no_sheet:
                print '    ' + xls_list[f] + ': ' + cert_no_file + \
                    ' cert_no in file name does not match cert_no within sheet ' + cert_no_sheet
            if cert_no_sheet not in cert_list:
                print '    ' + xls_list[f] + ': ' + cert_no_sheet + ' certificate not in DB'
'''   
db_conn.close()
'''
print '\nJob done.'

if __name__ == "__main__":
    main()
