# -*- coding: cp1252 -*-
'''+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
   This script is inspired by and modified from Tian Han's tillDB_data_loader script created for the import
   of data into the tillDB

   This script is coded to load and update the following 2 tables: 'data_sample'and 'data_ar' of the ARIS
   geochem staging database using the data provided in the staged xlsx files.
   
   Input
         1) Path to the directory containing the staged xlsx files
         2) Path to the aris geochem staginb database, ARIS_geochem_stage.accdb
         
  Output
         Data inserted into the above 2 tables in ARIS_geochem_stage.accdb
        
  Additional info
         1) Put all the staged xlsx files in a single directory. Data published in each publication
            have to be staged into a single staged xlsx file.
            
         2) In case a set of samples appeared in multiple publications, the corresponding staged xlsx
            file for each publication should be prepared separately with the same set of samples.
            The relationship between these samples and the corresponding publications is specified in
            'data_ar' table. To avaid duplication, these samples are saved only once in the
            'data_sample' table of the database.

         3) It is assumed that the data in these xlsx files are error free. All errors have
            to be addressed during the data screening process. So it is required to run
            'data_screener.py' (before execution of this script) to check whether data are properly
            staged. 
         
         6) All entries in the staged xlsx files are assumed to be in plain text. So make sure
            no subscript or superscript exists.
            
         7) This script only works when the input data are organized strictly following the exact
            format defined in the data staging.
            
         8) The current algorithm can not handle the following situations:

            c) Same samples republished using different 'sample_name'. For example, sample_A was originally
            published in 1988. It was re-published in 1990 as sample_B. In this case, both sample_A
            and sample_B will be recorded as 2 rows in 'data_sample' table and all their analytes included
            will be recorded in 'data_analyte' table. This causes redundency in both tables.

            All these problems are difficult to spot in data screening. They need to be examined visually in
            the final data products.
  Status
      Operational

  Developer
      Gabe Fortin
      
  Last update
      2017-06-07
  +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++'''
import os, sys, csv, pyodbc, ogr, osr
from openpyxl import load_workbook
#========================================== Sub-routines =================================================
#Get next row number of a given field and table
#Syntax: get_rownum(db_cursor, string, string) return int
def get_rownum(db_cur, fld_name, tab_name):
    db_cur.execute('select max(%s) from %s' %(fld_name, tab_name))

    max_val = db_cur.fetchone()
    if max_val[0] == None:
        return 1
    else:
        return int(max_val[0]) + 1
    
#============================================= Main routine =========================================

db_path = 'C:\\Project\\ARIS_Geochem_dev\\data\\ARIS_geochem_stage.accdb'
data_dir = 'C:\\Project\\ARIS_Geochem_dev\\data\\_AR Data Staging Location\\'

#Database connection
db_conn = pyodbc.connect('Driver={Microsoft Access Driver (*.mdb, *.accdb)};DBQ='+db_path)
cur = db_conn.cursor()

#Get next id values from tables: 'data_sample' and 'data_ar'
sample_id = get_rownum(cur, 'sample_id', 'data_sample')
ar_id = get_rownum(cur, 'ar_id', 'data_ar')

#Collect all xls file name under the specified directory
xls_list = os.listdir(data_dir)

#Loop through each file in the drectory
for f in range(len(xls_list)):
    xls_name = data_dir + xls_list[f]
    #extract the publication id (i.e. ARIS report number) from the file name
    ar_number = xls_list[f].partition('_')[0]

    #Open and exam each xls file
    wb = load_workbook(filename = xls_name)
    ws = wb[wb.sheetnames[0]]
    print xls_name + ' is being loaded ...'

    #Step through the data rows (values are all read in as strings)
    if (ws.cell(row = ws.max_row, column = 1)).value is None:
        reallastrow = ws.max_row - 1
    else:
        reallastrow = ws.max_row

    for r in range(2, reallastrow + 1):

        #-------------------------------------- Update the 2 relevent tables  -----------------------------------
        sample_name = str(ws.cell(row = r, column = 1).value).replace(' ', '')
        cur.execute("""select sample_id from data_sample where sample_name = ?""", sample_name)
        check_sample = cur.fetchone()
            
        #The current sample is already in 'data_sample' table (which won't be updated).
        if check_sample <> None:
            #----------------------------- Update 'data_ar' table if applicable ----------------------------
            cur.execute("""select count(*) from data_ar where (sample_id = ?) and (ar_number = ?)""",
                            check_sample[0], ar_number)
            rec_count = cur.fetchone()
            if rec_count[0] == 0:
                cur.execute("""insert into data_ar values (?, ?, ?)""", ar_id, ar_number, check_sample[0])
                cur.commit()
                ar_id = ar_id + 1

        #The current sample is not in 'data_sample' table (which is to be updated)
        else:
            sample_name = str(ws.cell(row = r, column = 1).value)

            station_name = str(ws.cell(row = r, column = 2).value)
            if (station_name == '' or station_name == 'None'):
                station_name = ''
                    
            sample_type = str(ws.cell(row = r, column = 3).value)

            sample_subtype = str(ws.cell(row = r, column = 4).value)
            if (sample_subtype == '' or sample_subtype == 'None'):
                sample_subtype = ''

            sample_depth = str(ws.cell(row = r, column = 5).value).replace(' ', '')
            if (sample_depth == '' or sample_depth == 'None'):
                sample_depth = ''

            if (ws.cell(row = r, column = 6).value) == None:
                sample_colour = None
            else:
                sample_colour = str((ws.cell(row = r, column = 6).value).replace(u'\xb1',"+/-"))
                if (sample_colour == '' or sample_colour == 'None'):
                    sample_colour = ''

            if (ws.cell(row = r, column = 7).value) == None:
                sample_desp = None
            else:
                sample_desp = str((ws.cell(row = r, column = 7).value).replace(u'\xb1',"+/-"))
                if (sample_desp == '' or sample_desp == 'None'):
                    sample_desp = ''

            duplicate = str(ws.cell(row = r, column = 8).value).replace(' ', '')
            if (duplicate == '' or duplicate == 'None'):
                duplicate = ''

            x_coord = str(ws.cell(row = r, column = 9).value).replace(' ', '')
            y_coord = str(ws.cell(row = r, column = 10).value).replace(' ', '')
            z_coord = str(ws.cell(row = r, column = 11).value).replace(' ', '')
            if z_coord == 'None':
                z_coord = None
            epsg_srid = str(ws.cell(row = r, column = 12).value).replace(' ', '')
            coord_conf = str(ws.cell(row = r, column = 13).value).replace(' ', '').upper()

            sample_date = str(ws.cell(row = r, column = 14).value)
            if (sample_date == '' or sample_date == 'None'):
                sample_date = None

            #Assemble a row of values to be written to 'data_sample' table
            sample_values = [sample_id, sample_name, station_name, sample_type, sample_subtype, sample_depth,
                                sample_colour, sample_desp, duplicate, x_coord, y_coord, z_coord, int(epsg_srid),
                                coord_conf, sample_date]

            #----------------------------- Add to the "data_sample" table ------------------------------------------
            cur.execute("""insert into data_sample values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""", sample_values)
            cur.commit()
            #---------------------------------- Update 'data_ar' table ----------------------------------------
            cur.execute("""select count(*) from data_ar where (sample_id = ?) and (ar_number = ?)""",
                        sample_id, ar_number)
            rec_count = cur.fetchone()
            if rec_count[0] == 0:
                cur.execute("""insert into data_ar values (?, ?, ?)""", ar_id, ar_number, sample_id)
                cur.commit()
                ar_id = ar_id + 1

            sample_id = sample_id + 1
            
db_conn.close()
print 'Job done!'
