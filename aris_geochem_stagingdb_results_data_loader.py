# -*- coding: cp1252 -*-
'''+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
   This script is inspired by and modified from Tian Han's tillDB_data_loader script created for the import
   of data into the tillDB

   This script is coded to load and update the following 2 tables: 'data_sample'and 'data_ar' of the ARIS
   geochem staging database using the data provided in the staged xlsx files.
   
   Input
         1) Path to the directory containing the staged xlsx files
         2) Path to the aris geochem staginb database, ARIS_geochem_stage.accdb
         
  Output
         Data inserted into the above 2 tables in ARIS_geochem_stage.accdb
        
  Additional info
         1) Put all the staged xlsx files in a single directory. Data published in each publication
            have to be staged into a single staged xlsx file.
            
         2) In case a set of samples appeared in multiple publications, the corresponding staged xlsx
            file for each publication should be prepared separately with the same set of samples.
            The relationship between these samples and the corresponding publications is specified in
            'data_ar' table. To avaid duplication, these samples are saved only once in the
            'data_sample' table of the database.

         3) It is assumed that the data in these xlsx files are error free. All errors have
            to be addressed during the data screening process. So it is required to run
            'data_screener.py' (before execution of this script) to check whether data are properly
            staged. 
         
         6) All entries in the staged xlsx files are assumed to be in plain text. So make sure
            no subscript or superscript exists.
            
         7) This script only works when the input data are organized strictly following the exact
            format defined in the data staging.
            
         8) The current algorithm can not handle the following situations:

            c) Same samples republished using different 'sample_name'. For example, sample_A was originally
            published in 1988. It was re-published in 1990 as sample_B. In this case, both sample_A
            and sample_B will be recorded as 2 rows in 'data_sample' table and all their analytes included
            will be recorded in 'data_analyte' table. This causes redundency in both tables.

            All these problems are difficult to spot in data screening. They need to be examined visually in
            the final data products.
  Status
      Operational

  Developer
      Gabe Fortin
      
  Last update
      2017-06-07
  +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++'''
import os, sys, csv, pyodbc, ogr, osr
from openpyxl import load_workbook


# ========================================== Sub-routines =================================================
# Get next row number of a given field and table
# Syntax: get_rownum(db_cursor, string, string) return int
def get_rownum(db_cur, fld_name, tab_name):
    db_cur.execute('select max(%s) from %s' % (fld_name, tab_name))

    max_val = db_cur.fetchone()
    if max_val[0] == None:
        return 1
    else:
        return int(max_val[0]) + 1


# Get unit_id with the given unit_name as defined in 'code_unit' table
# Syntax: get_unitid (db_cursor, list) return list
def get_unitid(db_cur, name_list):
    temp_id = []
    temp_name = []
    db_cur.execute("""select unit_id, name from code_unit""")
    all_records = db_cur.fetchall()
    for record in all_records:
        temp_id.append(record[0])
        temp_name.append(record[1])

    unitid_list = []
    for i in range(len(name_list)):
        name_list[i] = name_list[i].lower()

        # Standarize unit name
        if name_list[i].lower() == 'g':
            name_list[i] = 'g'

        # Validate
        if name_list[i] in temp_name:
            indx = temp_name.index(name_list[i])
            unitid_list.append(temp_id[indx])
        else:
            print 'Invalid unit: ' + name_list[i]
            sys.exit()

    return unitid_list

# Get method_id with the given method_abbr as defined in 'code_method' table
# Syntax: get_methodid (db_cursor, list) return list
def get_methodid(db_cur, name_list):
    temp_id = []
    temp_name = []


    methodid_list = []
    for i in range(len(name_list)):
        db_cur.execute("""select method_id from code_method where (method_abbr = ?)""", name_list[i])
        method_id = db_cur.fetchone()[0]
        methodid_list.append(method_id)

    return methodid_list


# ============================================= Main routine =========================================
def main():
    # File path
    db_path = 'C:\\Project\\ARIS_Geochem_dev\\data\\ARIS_geochem_stage.accdb'
    data_dir = 'C:\\Project\\ARIS_Geochem_dev\\data\\_AR Data Staging Results\\'

    # Database connection
    db_conn = pyodbc.connect('Driver={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=' + db_path)
    cur = db_conn.cursor()

    # Get next id values from tables ''data_analyte'
    analyte_id = get_rownum(cur, 'analyte_id', 'data_analyte')

    # Collect all xls file name under the specified directory
    xls_list = os.listdir(data_dir)

    # Loop through each file in the directory
    for f in range(len(xls_list)):
        xls_name = data_dir + xls_list[f]

        # Open and exam each xls file
        wb = load_workbook(filename=xls_name)
        ws = wb[wb.sheetnames[0]]
        print xls_name + ' is being loaded ...'

        #Get AR number from filename
        ar_number = xls_list[f].partition('_')[0]

        # Extract the top 4 rows from results xlsx
        analyte_list = list()
        unit_list = list()
        mdl_list = list()
        method_list = list()

        reallastcolumn = ws.max_column
        for i in range(ws.max_column, 3, -1):
            if (ws.cell(row=1, column=i)).value is None:
                reallastcolumn = i - 1
        for c in range(3, reallastcolumn + 1):
            analyte_list.append(str(ws.cell(row=1, column=c).value).replace(' ', ''))
            unit_list.append(str(ws.cell(row=2, column=c).value).replace(' ', ''))
            mdl_list.append(str(ws.cell(row=3, column=c).value).replace(' ', ''))
            method_list.append(str(ws.cell(row=4, column=c).value).replace(' ', ''))

        #Check that there is no data from this certificate in data_analyte
        certs_in_analyte = list()  # Build a list of method_id using current database values
        cur.execute("""select cert_no from vw_certs_in_data_analyte""")
        val_rows = cur.fetchall()
        for i in range(len(val_rows)):
            certs_in_analyte.append(str(val_rows[i][0]))

        impt_cert_no = (ws.cell(row=6, column=2)).value

        if impt_cert_no in certs_in_analyte:
            print 'Certificate already has data in the data_analyte table, it has likely already been imported.'
            sys.exit()

        #get cert_id from impt_cert_no
        cur.execute("""select cert_id from data_cert where (cert_no = ?)""", impt_cert_no)
        cert_id = cur.fetchone()[0]

        # Get 'unit_id' for the retrieved unit names
        unitid_list = get_unitid(cur, unit_list)

        methodid_list = get_methodid(cur, method_list)

        # Step through the remaining rows (values are all read in as strings)
        if (ws.cell(row=ws.max_row, column=1)).value is None:
            reallastrow = ws.max_row - 1
        else:
            reallastrow = ws.max_row
        for r in range(6, reallastrow + 1):

            #Find sample id based on AR number and sample_name
            cur.execute("""select sample_id from vw_ar_no_sampid_link where (ar_number = ?) and (sample_name = ?)""",
                        ar_number, str((ws.cell(row=r, column=1)).value))
            sample_id = cur.fetchone()[0]

            # Retrieve analyte values
            abundance = list()
            reallastcolumn = ws.max_column
            for i in range(ws.max_column, 3, -1):
                if (ws.cell(row=1, column=i)).value is None:
                    reallastcolumn = i - 1
            for c in range(3, reallastcolumn + 1):
                abundance.append(str(ws.cell(row=r, column=c).value).replace('<', '-'))

            # ----------------------------- Add to the "data_analyte" table -----------------------------------------
            for i in range(len(analyte_list)):
                cur.execute("""select count(*) from data_analyte where (analyte = ?) and (abundance = ?) and
                           (mdl = ?) and (unit_id = ?) and (method_id = ?) and  (sample_id = ?) and
                           (cert_id = ?)""", analyte_list[i], abundance[i], mdl_list[i], int(unitid_list[i]), \
                            int(methodid_list[i]), sample_id, cert_id)
                check_analyte = cur.fetchone()

                if (check_analyte[0] == 0) and (abundance[i] <> '') and (abundance[i] <> 'None'):
                    cur.execute("""insert into data_analyte values (?, ?, ?, ?, ?, ?, ?, ?)""", analyte_id, \
                                analyte_list[i], abundance[i], mdl_list[i], int(unitid_list[i]), \
                                int(methodid_list[i]), sample_id, cert_id)
                    cur.commit()
                    analyte_id = analyte_id + 1



#db_conn.close()
print 'Job done!'


if __name__ == "__main__":
    main()

