# -*- coding: cp1252 -*-
'''+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
   This script is inspired by and modified from Tian Han's tillDB_data_loader script created for the import
   of data into the tillDB

   This script is coded to load and update the following data_cert table of the ARIS
   geochem staging database using the data provided in the staged xlsx files.
   
   Input
         1) Path to the directory containing the staged xlsx files
         2) Path to the aris geochem staginb database, ARIS_geochem_stage.accdb
         
  Output
         Data inserted into the above table in ARIS_geochem_stage.accdb

  Additional info
         1) Put all the staged xlsx files in a single directory. Data published in each publication
            have to be staged into a single staged xlsx file.
            
         2) In case a set of samples appeared in multiple publications, the corresponding staged xlsx
            file for each publication should be prepared separately with the same set of samples.
            The relationship between these samples and the corresponding publications is specified in
            'data_ar' table. To avaid duplication, these samples are saved only once in the
            'data_sample' table of the database.

         3) It is assumed that the data in these xlsx files are error free. All errors have
            to be addressed during the data screening process. So it is required to run
            'data_screener.py' (before execution of this script) to check whether data are properly
            staged. 
         
         6) All entries in the staged xlsx files are assumed to be in plain text. So make sure
            no subscript or superscript exists.
            
         7) This script only works when the input data are organized strictly following the exact
            format defined in the data staging.
            
         8) The current algorithm can not handle the following situations:

            c) Same samples republished using different 'sample_name'. For example, sample_A was originally
            published in 1988. It was re-published in 1990 as sample_B. In this case, both sample_A
            and sample_B will be recorded as 2 rows in 'data_sample' table and all their analytes included
            will be recorded in 'data_analyte' table. This causes redundency in both tables.

            All these problems are difficult to spot in data screening. They need to be examined visually in
            the final data products.
  Status
      Operational

  Developer
      Gabe Fortin
      
  Last update
      2017-06-07
  +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++'''
import os, sys, csv, pyodbc, ogr, osr
from openpyxl import load_workbook
#========================================== Sub-routines =================================================
#Get next row number of a given field and table
#Syntax: get_rownum(db_cursor, string, string) return int
def get_rownum(db_cur, fld_name, tab_name):
    db_cur.execute('select max(%s) from %s' %(fld_name, tab_name))

    max_val = db_cur.fetchone()
    if max_val[0] == None:
        return 1
    else:
        return int(max_val[0]) + 1
    
#============================================= Main routine =========================================

db_path = 'C:\\Project\\ARIS_Geochem_dev\\data\\ARIS_geochem_stage.accdb'
data_dir = 'C:\\Project\\ARIS_Geochem_dev\\data\\_AR Data Staging Certificate\\'

#Database connection
db_conn = pyodbc.connect('Driver={Microsoft Access Driver (*.mdb, *.accdb)};DBQ='+db_path)
cur = db_conn.cursor()

#Collect all xls file name under the specified directory
xls_list = os.listdir(data_dir)

#Loop through each file in the drectory
for f in range(len(xls_list)):
    xls_name = data_dir + xls_list[f]

    # Get next id value from table: 'data_cert'
    cert_id = get_rownum(cur, 'cert_id', 'data_cert')

    #Open and exam each xls file
    wb = load_workbook(filename = xls_name)
    ws = wb[wb.sheetnames[0]]
    print xls_name + ' is being loaded ...'

    #Step through the data rows (values are all read in as strings)
    #-------------------------------------- Update the relevent table  -----------------------------------
    cert_no = str(ws.cell(row = 2, column = 1).value).replace(' ', '')
    cur.execute("""select cert_id from data_cert where cert_id = ?""", cert_no)
    check_cert = cur.fetchone()
            
    #The current certificate is already in 'data_cert' table (which won't be updated).
    if check_cert <> None:
        print xls_list[f] + ' cert_no: ' + cert_no + ' is already in the database and will not be re-imported'
    #The current certificate is not in 'data_cert' table (which is to be updated)
    else:
        cert_date = str(ws.cell(row = 2, column = 2).value)
        lab_id = str(ws.cell(row=2, column=3).value).replace(' ', '')
        prep_id = str(ws.cell(row=2, column=4).value).replace(' ', '')

        #Assemble a row of values to be written to 'data_cert' table
        cert_values = [cert_id, cert_no, cert_date, int(lab_id), int(prep_id), '']

        #----------------------------- Add to the "data_cert" table ------------------------------------------
        cur.execute("""insert into data_cert values (?, ?, ?, ?, ?, ?)""", cert_values)
        cur.commit()

            
db_conn.close()
print 'Job done!'
