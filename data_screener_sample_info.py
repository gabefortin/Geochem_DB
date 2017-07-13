# -*- coding: cp1252 -*-
'''++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
   This script is modified from Tian Han's data_screener.py script. It has been adapted to
   check sample location information for the ARIS Geochem data compilation project.
   
   This script is coded to identify and report problems found in the staged geochem sample
   location data in xlsx files before loading them loaded into the ARIS Geochem database.
   Please beware that this script only identifies and reports problems. The problems identified
   need to be fixed manually.
   
   The file names for the xlsx sheets are used to code the ARIS numbers into the database, so
   the names are important. All files should be name XXXXX_loc.xlsx where XXXXX is the ARIS
   report number.
   
   Input
         1) Path to the directory where the staged xlsx files reside
         2) aris_geochem Database path (this is not currently necessary as of June 1, 2017
         but may be used at a later date).
         3) Some hardcoded parameters 
         
   Output
         Display identified problems

   Operation note
         1) The script has many independent code-blocks targeting different types of errors as listed
            below. Though each code-block can be executed randomly, it is strongly recommended to
            run them sequentially in the given order, one at a time. For example, it should start with
            checking format related problems in the xlsx files by commenting out all other code-blocks
            which are for checking other types of errors. Once the format-related errors are identified
            and fixed, the format-checking code-block should be commented out and then move on to the
            "analyte name checking block" ...

         2) Duplicate samples should be given different "sample_code". Re-published samples should
            have the same 'sample_code' across all publications. Sections 15 and 16 are designed to
            ensure if the above requirements are met.
            
    This script is able to check the following:
         1)  Check the format of the spreadsheet to ensure that all necessary columns are present, properly named and in
            the right order.
         2) Check that x, y, z coordinates and epsg codes are present and are in fact numeric. Check that z values, when
            present are between 0 and 3000 metres.
         3) Check that values entered in the sample_type column match the pre-determined options. The script will 
            automatically turn any upper case character to lower case.
         4)  Check that values in sample_subtype match accepted values. This column is case sensitive.
         5)  Check that values entered in coord_conf match accepted values.
         6)  Confirm that points actually plot within the province of BC by computing NAD83 Lat Long on the fly and
            comparing to a shapefile with the BC boundary.
        7)  Check that all dates under sample_date are proper dates.
         
    
  Status
      In Development

  Future improvements
        The algorithms used in 16) and 17) above generate too many false alarms. They are not helpful
        enough to identify duplicate samples.

  Developers
      T. Han, G. Fortin
      
  Last update
      2017-05-21
  +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++'''
import os, ogr, osr, datetime, openpyxl, pyodbc
from openpyxl import load_workbook
from dateutil.parser import parse

# This function is to check if a given string can be converted to a decimal number
# Syntax: is_number(string) return logic
def is_number(n):
    try:
        float(n)
        return True
    except ValueError:
        return False
    
# This function is to check if a given variable can be converted to a string
# Syntax: is_number(variable) return logic
def is_string(s):
    try:
        str(s)
        return True
    except ValueError:
        return False
    
#Project input coordinates to NAD83 lat. and long (EPSG_SRID = 4269)
#Syntax: project2nad83(float, float, int) returns [nad83_long, nad83_lat]
def project2nad83 (x_coord, y_coord, source_epsg):

    #Create a point geometry using the given coordinates
    point = ogr.Geometry(ogr.wkbPoint)
    point.AddPoint(x_coord, y_coord)

    #Specify source spatial reference and projection
    source_ref = osr.SpatialReference()
    source_ref.ImportFromEPSG(source_epsg)

    #Define target spatial reference and projection based on input
    target_ref = osr.SpatialReference()
    target_ref.ImportFromEPSG(4269)
        
    #Perform transformation
    transform = osr.CoordinateTransformation(source_ref, target_ref)
    point.Transform(transform)
    
    return [str(point.GetX(0)), str(point.GetY(0))]
    
def main():   
    #File path
    data_dir = 'C:\\Project\\ARIS_Geochem_dev\\data_testing\\_AR Data Staging Location\\'
    chkrpt_nm = 'C:\\Project\\ARIS_Geochem_dev\\data_testing\\checkreports\\SampleInfoCheckReport_' + \
        datetime.datetime.now().strftime("%Y_%m_%d_%H_%M") + '.xlsx'


    #Minimum long. and lat. difference between sample locations in degree
    min_lon = 0.001
    min_lat = 0.001
    
    #Collect year sub-directories under given "data_dir" 
    xls_list = os.listdir(data_dir)

    #List of allowed EPSG codes
    epsg_list = ['2955','3156','3157','4269','4326','26709','26710']

    #List of allowed sample_types
    samptype_list = ['soil','soil-mmi','silt','stream sediment','till','moss mat']

    #List of allowed sample_subtypes
    subtype_list = ['A Horizon','B Horizon','C Horizon','Ah Horizon','A-B Horizons','B-C Horizons','A-C Horizons',
                    'silt','pit','trench','pan concentrate']

    #create check report xlsx file and 'open' it
    chkwb = openpyxl.Workbook()
    chkwb.remove_sheet(chkwb.get_sheet_by_name('Sheet'))
    chkws = chkwb.create_sheet(title='CheckResults')
    chkws.cell(row=1, column=1).value = 'File Name'
    chkws.cell(row=1, column=2).value = 'Check Type'
    chkws.cell(row=1, column=3).value = 'Problem'
    chkws.cell(row=1, column=4).value = 'Row'
    chkws.cell(row=1, column=5).value = 'Column'

    #function to write problems to check report xlsx
    def write_rpt_err(file,check,problem,row,column):
        chkws.cell(row=chkws.max_row + 1, column=1).value = file
        chkws.cell(row=chkws.max_row, column=2).value = check
        chkws.cell(row=chkws.max_row, column=3).value = problem
        chkws.cell(row=chkws.max_row, column=4).value = row
        chkws.cell(row=chkws.max_row, column=5).value = column

    #====================================== 1. Check xls file format ======================================
    print '1. Examine file format ...'

    #Loop through xls file
    for f in range(len(xls_list)):
        xls_name = data_dir + xls_list[f]
            
        #Open and exam each xls file
        wb = load_workbook(filename = xls_name)
        ws = wb[wb.sheetnames[0]]

        #get total number of rows in the spreadsheet, this code avoids counting empty rows as the last row and
        #reduces the number of false errors.
        if (ws.cell(row = ws.max_row, column = 1)).value is None:
            reallastrow = ws.max_row - 1
        else:
            reallastrow = ws.max_row

        if (str(ws.cell(row = 1, column = 1).value)).lower() <> 'sample_name':
            write_rpt_err(xls_list[f],'File Format','Wrong sample_name column header',1,1)
        if (str(ws.cell(row = 1, column = 2).value)).lower() <> 'station_name':
            write_rpt_err(xls_list[f], 'File Format', 'Wrong station_name column header', 1, 2)
        if (str(ws.cell(row = 1, column = 3).value)).lower() <> 'sample_type':
            write_rpt_err(xls_list[f], 'File Format', 'Wrong sample_type column header', 1, 3)
        if (str(ws.cell(row = 1, column = 4).value)).lower() <> 'sample_subtype':
            write_rpt_err(xls_list[f], 'File Format', 'Wrong sample_subtype column header', 1, 4)
        if (str(ws.cell(row = 1, column = 5).value)).lower() <> 'sample_depth':
            write_rpt_err(xls_list[f], 'File Format', 'Wrong sample_depth column header', 1, 5)
        if (str(ws.cell(row = 1, column = 6).value)).lower() <> 'sample_colour':
            pwrite_rpt_err(xls_list[f],'File Format','Wrong sample_colour column header',1,6)
        if (str(ws.cell(row = 1, column = 7).value)).lower() <> 'sample_desp':
            write_rpt_err(xls_list[f], 'File Format', 'Wrong sample_desp column header', 1, 7)
        if (str(ws.cell(row = 1, column = 8).value)).lower() <> 'duplicate':
            write_rpt_err(xls_list[f], 'File Format', 'Wrong duplicate column header', 1, 8)
        if (str(ws.cell(row = 1, column = 9).value)).lower() <> 'x_coord':
            write_rpt_err(xls_list[f], 'File Format', 'Wrong x_coord column header', 1, 9)
        if (str(ws.cell(row = 1, column = 10).value)).lower() <> 'y_coord':
            write_rpt_err(xls_list[f], 'File Format', 'Wrong y_coord column header', 1, 10)
        if (str(ws.cell(row = 1, column = 11).value)).lower() <> 'z_coord':
            write_rpt_err(xls_list[f], 'File Format', 'Wrong z_coord column header', 1, 11)
        if (str(ws.cell(row = 1, column = 12).value)).lower() <> 'epsg_srid':
            write_rpt_err(xls_list[f], 'File Format', 'Wrong epsg_srid column header', 1, 12)
        if (str(ws.cell(row = 1, column = 13).value)).lower() <> 'coord_conf':
            write_rpt_err(xls_list[f], 'File Format', 'Wrong coord_conf column header', 1, 13)
        if (str(ws.cell(row = 1, column = 14).value)).lower() <> 'sample_date':
            write_rpt_err(xls_list[f], 'File Format', 'Wrong sample_date column header', 1, 1)

    chkwb.save(chkrpt_nm) #save results to the xlsx report

    #============================ 2. Check x_coord, y_coord, z_coord, and epsg_srid ==========================
    print '\n2.Examine x-coord, y-coord, z-coord and epsg_srid ...'

    #Loop through xls files
    for f in range(len(xls_list)):
        xls_name = data_dir + xls_list[f]
            
        #Open and exam each xls file
        wb = load_workbook(filename = xls_name)
        ws = wb[wb.sheetnames[0]]
        if (ws.cell(row = ws.max_row, column = 1)).value is None:
            reallastrow = ws.max_row - 1
        else:
            reallastrow = ws.max_row

        for r in range(2, reallastrow + 1):
            
            x_cell = str(ws.cell(row = r, column = 9).value).replace(' ', '')
            if not is_number(x_cell):
                write_rpt_err(xls_list[f], 'Coordinates', 'x_coord is not a number', str(r), 9)

            y_cell = str(ws.cell(row = r, column = 10).value).replace(' ', '')
            if not is_number(y_cell):
                write_rpt_err(xls_list[f], 'Coordinates', 'y_coord is not a number', str(r), 10)

            z_cell = str(ws.cell(row = r, column = 11).value).replace(' ', '')
            if not ((z_cell == '') or (z_cell == 'None')):
                if (not is_number(z_cell)):
                    write_rpt_err(xls_list[f], 'Coordinates', 'z_coord is not a number', str(r), 11)
                if float(z_cell) < 0 or float(z_cell) > 3000:
                    write_rpt_err(xls_list[f], 'Coordinates', 'z_coord is out of range (i.e. not between 0 and 3000m)' \
                                  , str(r), 11)
            
            epsg_cell = str(ws.cell(row = r, column = 12).value).replace(' ', '')
            if not is_number(epsg_cell):
                write_rpt_err(xls_list[f], 'Coordinates', 'epsg_srid is not a number', str(r), 12)
            else:
                if epsg_cell not in epsg_list:
                    write_rpt_err(xls_list[f], 'Coordinates', 'epsg_srid not in list', str(r), 12)

    chkwb.save(chkrpt_nm) #save results to the xlsx report

    # ============================ 3. Check sample_type ==========================
    print '\n3.Examine sample_type ...'

    # Loop through xls files
    for f in range(len(xls_list)):
        xls_name = data_dir + xls_list[f]

        # Open and exam each xls file
        wb = load_workbook(filename=xls_name)
        ws = wb[wb.sheetnames[0]]
        if (ws.cell(row=ws.max_row, column=1)).value is None:
            reallastrow = ws.max_row - 1
        else:
            reallastrow = ws.max_row

        #if sample_type not all lowercase, convert and save to xlsx
        for r in range(2, reallastrow + 1):
            samptype_cell = str(ws.cell(row=r, column=3).value)
            if samptype_cell != str.lower(samptype_cell):
                ws.cell(row=r, column=3).value = str.lower(samptype_cell)
                wb.save(xls_name)

        #check that sample_type matches an option in list
        for r in range(2, reallastrow + 1):
            samptype_cell = str(ws.cell(row=r, column=3).value)
            if samptype_cell not in samptype_list:
                write_rpt_err(xls_list[f], 'Sample Type', 'sample_type not in list', str(r), 3)

    chkwb.save(chkrpt_nm)  # save results to the xlsx report

    #============================ 4. Check sample_subtype ==========================
    print '\n4.Examine sample subtype ...'

    #Loop through xls files
    for f in range(len(xls_list)):
        xls_name = data_dir + xls_list[f]

        #Open and exam each xls file
        wb = load_workbook(filename = xls_name)
        ws = wb[wb.sheetnames[0]]
        if (ws.cell(row = ws.max_row, column = 1)).value is None:
            reallastrow = ws.max_row - 1
        else:
            reallastrow = ws.max_row

        for r in range(2, reallastrow + 1):

            subtype_cell = str(ws.cell(row = r, column = 4).value)
            if not subtype_cell == 'None':
                if subtype_cell not in subtype_list:
                    write_rpt_err(xls_list[f], 'Sample Subtype', 'sample_subtype not in list', str(r), 4)

    chkwb.save(chkrpt_nm)  # save results to the xlsx report

    #====================================== 5. Check coord_conf =====================================
    print '\n5.Examine Coord_Conf ...'

    #Loop through xls file under year directory
    for f in range(len(xls_list)):
        xls_name = data_dir + xls_list[f]
            
        #Open and exam each xls file
        wb = load_workbook(filename = xls_name)
        ws = wb[wb.sheetnames[0]]
        if (ws.cell(row = ws.max_row, column = 1)).value is None:
            reallastrow = ws.max_row - 1
        else:
            reallastrow = ws.max_row

        for r in range(2, reallastrow + 1):
            work_cell = (str(ws.cell(row = r, column = 13).value).replace(' ', '')).lower()
            if (work_cell <> 'l') and (work_cell <> 'm') and (work_cell <> 'h'):
                write_rpt_err(xls_list[f], 'Coordinate Confidence', 'coord_conf not in list (l,m,h)', str(r), 13)

    chkwb.save(chkrpt_nm)  # save results to the xlsx report

    #================== 6. Check points are within BC and within 10km of ARIS point=============
    print '\n6.Examine sample locations to ensure they fall in BC  ...'

    #Looking for samples that do not fall in BC to identify possible coordinate issues

    # load the shape file as a layer
    drv    = ogr.GetDriverByName('ESRI Shapefile')
    ds_in  = drv.Open("prov_ab_p_geo83_e.shp")
    lyr_in = ds_in.GetLayer(0)

    # field index for which i want the data extracted
    # ("satreg2" was what i was looking for)
    idx_reg = lyr_in.GetLayerDefn().GetFieldIndex("NAME")

    #Loop through the xls files
    xls_sample = list()  #Build a list for sample_code in xlsx files
    xls_xc = list()      #Build a list for x_coord in xlsx files
    xls_yc = list()      #Build a list for y_coord in xlsx files
    xls_file = list()    #Build a list for xlsx file names
    xls_row = list()
    xls_loccheck = list()
    for f in range(len(xls_list)):
        xls_name = data_dir + xls_list[f]

        #Open and exam each xls file
        wb = load_workbook(filename = xls_name)
        ws = wb[wb.sheetnames[0]]

        if (ws.cell(row = ws.max_row, column = 1)).value is None:
            reallastrow = ws.max_row - 1
        else:
            reallastrow = ws.max_row


        for r in range(2, reallastrow + 1):
            cell_sample = str(ws.cell(row = r, column = 1).value).replace(' ', '')
            cell_xc = str(ws.cell(row = r, column = 9).value).replace(' ', '')
            cell_yc = str(ws.cell(row = r, column = 10).value).replace(' ', '')
            cell_epsg = str(ws.cell(row = r, column = 12).value).replace(' ', '')
            cell_row = str(r)

            xls_sample.append(cell_sample)

            if cell_epsg <> '4269': #NAD83 geographic
                [cell_xc, cell_yc] = project2nad83 (float(cell_xc), float(cell_yc), int(cell_epsg))
            xls_xc.append(cell_xc)
            xls_yc.append(cell_yc)
            xls_file.append(xls_name)
            xls_row.append(cell_row)


    #Loop through to check each sample coord
    for i in range(len(xls_sample)):
        #Examine sample locations
        coordcheck = 2

        # create point geometry
        pt = ogr.Geometry(ogr.wkbPoint)
        pt.SetPoint_2D(0, float(xls_xc[i]), float(xls_yc[i]))
        lyr_in.SetSpatialFilter(pt)

        # go over all the polygons in the layer see if one include the point
        for feat_in in lyr_in:
            # roughly subsets features, instead of go over everything
            ply = feat_in.GetGeometryRef()
            # test
            if pt.Within(ply):
                coordcheck = 1
            else:
                coordcheck = 0

        if coordcheck != 1:
            print xls_file[i] + ', Row: ' + xls_row[i] + ', Sample: ' + xls_sample[i]+ ': Location not in BC'
            write_rpt_err(xls_file[i], 'Location', 'Location not in BC', xls_row[i], '')

    chkwb.save(chkrpt_nm)  # save results to the xlsx report

    #----------------------------------7.Check date validity-----------------------------------------------------
    print '\n7.Examine sample dates ...'

    #Loop through xls files
    for f in range(len(xls_list)):
        xls_name = data_dir + xls_list[f]

        #Open and exam each xls file
        wb = load_workbook(filename = xls_name)
        ws = wb[wb.sheetnames[0]]
        if (ws.cell(row = ws.max_row, column = 1)).value is None:
            reallastrow = ws.max_row - 1
        else:
            reallastrow = ws.max_row

        for r in range(2, reallastrow + 1):

            date_cell = str(ws.cell(row = r, column = 14).value)

            if not ((date_cell == '') or (date_cell == 'None')):
                if not (date_cell.isupper() or (date_cell.islower())):
                    try:
                        parse(date_cell)
                    except ValueError:
                        write_rpt_err(xls_list[f], 'Date', str(ws.cell(row=r, column=14).value) + \
                                      'is not a proper date', str(r), 14)
                else:
                    write_rpt_err(xls_list[f], 'Date', str(ws.cell(row=r, column=14).value) + \
                                  'is not a proper date', str(r), 14)

    chkwb.save(chkrpt_nm)  # save results to the xlsx report

    #================== 8. Check points are within 10km of ARIS point=============
    print '\n8. Check points are within 10km of ARIS point  ...'

    #This check is currently very slow. Needs to be improved.... can be ignored if time is an issue.

    #Looking for samples that do not fall in BC to identify possible coordinate issues

    # load the shape file as a layer
    drv    = ogr.GetDriverByName('ESRI Shapefile')
    ds_in  = drv.Open("aris_10km_buffer.shp")
    lyr_in = ds_in.GetLayer(0)

    # field index for which i want the data extracted
    # ("satreg2" was what i was looking for)
    idx_reg = lyr_in.GetLayerDefn().GetFieldIndex("asses_num")

    #Loop through the xls files
    xls_sample = list()  #Build a list for sample_code in xlsx files
    xls_xc = list()      #Build a list for x_coord in xlsx files
    xls_yc = list()      #Build a list for y_coord in xlsx files
    xls_file = list()    #Build a list for xlsx file names
    xls_row = list()
    xls_loccheck = list()
    for f in range(len(xls_list)):
        xls_name = data_dir + xls_list[f]

        #Open and exam each xls file
        wb = load_workbook(filename = xls_name)
        ws = wb[wb.sheetnames[0]]

        if (ws.cell(row = ws.max_row, column = 1)).value is None:
            reallastrow = ws.max_row - 1
        else:
            reallastrow = ws.max_row


        for r in range(2, reallastrow + 1):
            cell_sample = str(ws.cell(row = r, column = 1).value).replace(' ', '')
            cell_xc = str(ws.cell(row = r, column = 9).value).replace(' ', '')
            cell_yc = str(ws.cell(row = r, column = 10).value).replace(' ', '')
            cell_epsg = str(ws.cell(row = r, column = 12).value).replace(' ', '')
            cell_row = str(r)

            xls_sample.append(cell_sample)

            if cell_epsg <> '4269': #NAD83 geographic
                [cell_xc, cell_yc] = project2nad83 (float(cell_xc), float(cell_yc), int(cell_epsg))
            xls_xc.append(cell_xc)
            xls_yc.append(cell_yc)
            xls_file.append(xls_name)
            xls_row.append(cell_row)


    #Loop through to check each sample coord
    for i in range(len(xls_sample)):
        #Examine sample locations

        # create point geometry
        pt = ogr.Geometry(ogr.wkbPoint)
        pt.SetPoint_2D(0, float(xls_xc[i]), float(xls_yc[i]))
        lyr_in.SetSpatialFilter(pt)

        # go over all the polygons in the layer see if one include the point
        for feat_in in lyr_in:
            if feat_in.GetFieldAsString("asses_num") == xls_list[f].partition('_')[0]:
                # roughly subsets features, instead of go over everything
                ply = feat_in.GetGeometryRef()
                # test
                if not pt.Within(ply):
                    print 'sample is within 10k of aris'
                    write_rpt_err(xls_file[i], 'Location', 'Location not within 10km of ARIS report location', \
                                  xls_row[i], '')
            else:
                write_rpt_err(xls_file[i], 'Location', 'No ARIS record was found for this report', \
                              xls_row[i], '')

    chkwb.save(chkrpt_nm)  # save results to the xlsx report

    print '\nJob done.'
    
if __name__ == "__main__":
    main()
