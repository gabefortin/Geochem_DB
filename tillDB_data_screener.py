# -*- coding: cp1252 -*-
'''+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
   This script is coded to identify and report problems found in the staged till data in xlsx files
   before loading them loaded into the tillDB database. Please beware that this script only identifies
   and reports problems. The problems identified need to be fixed manually.
   
   Input
         1) Path to the directory where the staged xlsx files reside
         2) tillDB Database path
         3) Some hardcoded parameteres 
         
   Output
         Display identified problems

   Operation note
         1) The script has many independent code-blocks targeting different types of errors as listed
            below. Though each code-block can be executed randomly, it is strongly recommended to
            run them sequentially in the given order, one at a time. For example, it should start with
            checking format related problems in the xlsx files by commenting out all other code-blocks
            which are for checking other types of errors. Once the format-related errors are identified
            and fixed, the format-checking code-block should be commented out and then move on to the
            "analyte name checking block" ...

         2) Duplicate samples should be given different "sample_code". Re-published samples should
            have the same 'sample_code' across all publications. Sections 15 and 16 are designed to
            ensure if the above requirements are met.
            
    This script is able to check the following:
         1)  if format of staged xlsx files is correct;
         2)  if analyte name are appropriate;
         3)  if analyte value unit exists in database;
         4)  if detection limit value present in xlsx files;
             Be noted that blank detection limit is ok. This entry is optional.
             
         5)  if analytical "method_id" exists in database;
         6)  if "lab_id" exists in database;
         7)  if "size_fraction" is blank, which is not allowed;
         8)  if duplicate columns exist within xlsx files; 
             It should be noted that duplicate columns may exist. For example, an analyte was initally
             and re-analyzed using the same method by the same lab. It often happens to Au which results
             in 2 Au columns. Extra caution should be exersized when dealing with these cases. 

         9)  if duplicate or empty samples exist within each xlsx file;
         10) if invalid x_coord, y_coord, z_coord, and EPSG_SRID values exist in xlsx files;
         11) if xlsx file name corresponds to "pub_issue" in xlsx files; 
         12) if "coord_conf" values are appropriate in xlsx files;
         13) if analyte value units are appropriate for some method-dependent analytes;
         14) if method-specific analytes use the right methods;
         15) if certain analyte values are appropriate

         16) if duplicate samples exist among the samples in the xlsx files.
             Duplicate-sample flag will be triggered if the following 2 conditions are met simultaneously:
             a) closely located;
             b) similar sample names.

             Be noted that 1) the flagged samples may be those that are re-published; 2) this algorithm
             doesn't work if 2 samples are named very differently through they are true duplicates; and
             3) the flagged samples are not necessarily true duplicates. They need to examined carefully.
             
         17) if duplciate samples exist among the samples in the xlsx files and those in the database.
             Duplicate-sample flag will be triggered if the following 2 conditions are met simultaneously:
             a) closely located;
             b) similar sample names.

             Be noted that 1) the flagged samples may be those that are re-published; 2)this algorithm
             doesn't work if 2 samples are named very differently through they are true duplicates; and
             3) the flagged samples are not necessarily true duplicates. They need to examined carefully.
    
  Status
      Operational

  Furture improment
        The algorithms used in 16) and 17) above generate too many false alarms. They are not helpful
        enough to identify duplicate samples.

  Developer
      T. Han
      
  Last update
      2016-12-15
  +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++'''
import os, sys, csv, pyodbc, ogr, osr
from openpyxl import load_workbook

# This function is to check if a given string can be converted to a decimal number
# Syntax: is_number(string) return logic
def is_number(n):
    try:
        float(n)
        return True
    except ValueError:
        return False
    
# This function is to check if a given variable can be converted to a string
# Syntax: is_number(variable) return logic
def is_string(s):
    try:
        str(s)
        return True
    except ValueError:
        return False
    
#Project input coordinates to NAD83 lat. and long (EPSG_SRID = 4269)
#Syntax: project2nad83(float, float, int) returns [nad83_long, nad83_lat]
def project2nad83 (x_coord, y_coord, source_epsg):

    #Create a point geometry using the given coordinates
    point = ogr.Geometry(ogr.wkbPoint)
    point.AddPoint(x_coord, y_coord)

    #Specify source spatial reference and projection
    source_ref = osr.SpatialReference()
    source_ref.ImportFromEPSG(source_epsg)

    #Define target spatial reference and projection based on input
    target_ref = osr.SpatialReference()
    target_ref.ImportFromEPSG(4269)
        
    #Perform transformation
    transform = osr.CoordinateTransformation(source_ref, target_ref)
    point.Transform(transform)
    
    return [str(point.GetX(0)), str(point.GetY(0))]
    
def main():   
    #File path
    db_path = 'C:\\Project\\TillDB\\data\\tillDB_curr.accdb'
    data_dir = 'C:\\Project\\TillDB\\data\\uploaded\\'

    #Database connection
    db_conn = pyodbc.connect('Driver={Microsoft Access Driver (*.mdb, *.accdb)};DBQ='+db_path)
    cur = db_conn.cursor()

    #Standard list of analyte elements (provided by Alexei).
    element_list = ['Si','Al','Ca','Fe','K','Mg','Na','P','S(T)','SO3','CO2','Sn','Sr','SiO2',\
                    'TiO2', 'Al2O3','Fe2O3(T)','FeO(T)','FeO','MnO','MgO','CaO','Na2O','K2O','P2O5',\
                    'V2O5','Cr2O3','CoO','NiO','CuO','ZnO','As2O3','SrO','Y2O3','ZrO2','Nb2O5','MoO2',\
                    'SnO','Sb2O3','BaO','TR2O3','La2O3','Lu2O3','Ta2O5','Ti','WO2','PbO','Bi2O3','H2O+',\
                    'H2O-','LOI','Total','Cl','Au','Pd','Pt','Re','Os','Ir','Ru','Rh','Hg','Ag', \
                    'Wt','As','B','Ba','Be','Bi','Br','Cd','Ce','Co','Cr','Cs','Cu','Dy','Er','Eu',\
                    'Ga','Gd','Ge','Hf','Ho','In','La','Li','Lu','Mo','Mn','Nb','Nd','Ni','Pb','Pr',\
                    'Rb','Sb','Sc','Se','Sm','Ta','Tb','Te','Th','Tl','Tm','U','V','W','Y','Yb',\
                    'Zn','Zr', 'F', 'C(T)', 'C(org)', 'C(inorg)', 'H2O(T)', 'N', 'CaF2']

    #Standard list of unit corresponding to the analyte above (provided by Alexei)
    standard_unit = ['%', '%', '%', '%', '%', '%', '%', '%', '%', '%', '%', 'ppm', 'ppm', '%', \
                     '%', '%', '%', '%', '%', '%', '%', '%', '%', '%', '%', \
                     '%', '%', '%', '%', '%', '%', '%', '%', '%', '%', '%', '%', \
                     '%', '%', '%', '%', '%', '%', '%', '%', '%', '%', '%', '%', \
                     '%', '%', '%', '%', 'ppb', 'ppb', 'ppb', 'ppb', 'ppb', 'ppb', 'ppb', 'ppb', 'ppb', 'ppm', \
                     'g', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', \
                     'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', \
                     'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', 'ppm', \
                     'ppm', 'ppm', 'ppm', '%', '%', '%', '%', 'ppm', '%']

    #List of analytes who units are method-dependent
    dependent_analyte = ['Ag', 'Au', 'Hg', 'Sn', 'Sr']
    dependent_method =  ['AIP', 'MIP', 'INA', 'INA', 'INA']
    dependent_unit =    ['ppm', 'ppb', 'ppb', 'ppm', 'ppm']
    
    #Standard_list for analytes without detection limit
    nolimit_list = ['Total', 'Wt']

    #List of fixed combo between analyte and method
    fix_analyte = ['LOI', 'H2O-', 'H2O(T)', 'C(T)', 'F', 'FeO', 'CO2', 'C(org)']
    fix_method =  ['7', '7', '11', '11', '20', '21', '11', '11']

    #Minimum long. and lat. difference between sample locations in degree
    min_lon = 0.001
    min_lat = 0.001
    
    #Collect year sub-directories under given "data_dir" 
    xls_list = os.listdir(data_dir)
    #====================================== 1. Check xls file format ======================================
    print '1. Examine file format ...'
    #Loop through xls file
    for f in range(len(xls_list)):
        xls_name = data_dir + xls_list[f]
            
        #Open and exam each xls file
        wb = load_workbook(filename = xls_name)
        ws = wb[wb.sheetnames[0]]
        if (str(ws.cell(row = 7, column = 1).value)).lower() <> 'sample_name':
            print '    ' + xls_name + ': Wrong Sample_Name column header'
        if (str(ws.cell(row = 7, column = 2).value)).lower() <> 'sample_code':
            print '    ' + xls_name + ': Wrong Sample_Code column header'
        if (str(ws.cell(row = 7, column = 3).value)).lower() <> 'sample_type':
            print '    ' + xls_name + ': Wrong Sample_Name column header'
        if (str(ws.cell(row = 7, column = 4).value)).lower() <> 'depth':
            print '    ' + xls_name + ': Wrong Depth column header'
        if (str(ws.cell(row = 7, column = 5).value)).lower() <> 'duplicate':
            print '    ' + xls_name + ': Wrong Duplicate column header'
        if (str(ws.cell(row = 7, column = 6).value)).lower() <> 'borehole':
            print '    ' + xls_name + ': Wrong Borehole column header'
        if (str(ws.cell(row = 7, column = 7).value)).lower() <> 'core_top':
            print '    ' + xls_name + ': Wrong Core_Top column header'
        if (str(ws.cell(row = 7, column = 8).value)).lower() <> 'core_bottom':
            print '    ' + xls_name + ': Wrong Core_Botton column header'
        if (str(ws.cell(row = 7, column = 9).value)).lower() <> 'azimuth':
            print '    ' + xls_name + ': Wrong Azimuth column header'
        if (str(ws.cell(row = 7, column = 10).value)).lower() <> 'dip':
            print '    ' + xls_name + ': Wrong Dip column header'
        if (str(ws.cell(row = 7, column = 11).value)).lower() <> 'drill_type':
            print '    ' + xls_name + ': Wrong Drill_Type column header'
        if (str(ws.cell(row = 7, column = 12).value)).lower() <> 'material_type':
            print '    ' + xls_name + ': Wrong Material_Type column header'
        if (str(ws.cell(row = 7, column = 13).value)).lower() <> 'sample_desc':
            print '    ' + xls_name + ': Wrong Sample_Desc column header'
        if (str(ws.cell(row = 7, column = 14).value)).lower() <> 'x-coord':
            print '    ' + xls_name + ': Wrong X-Coord column header'
        if (str(ws.cell(row = 7, column = 15).value)).lower() <> 'y-coord':
            print '    ' + xls_name + ': Wrong Y-Coord column header'
        if (str(ws.cell(row = 7, column = 16).value)).lower() <> 'z-coord':
            print '    ' + xls_name + ': Wrong Z-Coord column header'
        if (str(ws.cell(row = 7, column = 17).value)).lower() <> 'epsg_srid':
            print '    ' + xls_name + ': Wrong EPSG_SRID column header'
        if (str(ws.cell(row = 7, column = 18).value)).lower() <> 'pub_issue':
            print '    ' + xls_name + ': Wrong Pub_Issue column header'
        if (str(ws.cell(row = 7, column = 19).value)).lower() <> 'coord_conf':
            print '    ' + xls_name + ': Wrong Coord_Conf column header'
                
        if (str(ws.cell(row = 1, column = 19).value)).lower() <> 'analyte':
            print '    ' + xls_name + ': Wrong Analyte row header'
        if (str(ws.cell(row = 2, column = 19).value)).lower() <> 'unit':
            print '    ' + xls_name + ': Wrong Unit row header'
        if (str(ws.cell(row = 3, column = 19).value)).lower() <> 'd_limit':
            print '    ' + xls_name + ': Wrong D_Limit row header'
        if (str(ws.cell(row = 4, column = 19).value)).lower() <> 'method_id':
            print '    ' + xls_name + ': Wrong Method_ID row header'
        if (str(ws.cell(row = 5, column = 19).value)).lower() <> 'lab_id':
            print '    ' + xls_name + ': Wrong Lab_ID row header'
        if (str(ws.cell(row = 6, column = 19).value)).lower() <> 'size_fraction':
            print '    ' + xls_name + ': Wrong Size_Fraction row header' 
    #===================================== 2. Check analyte name =======================================
    print '\n2. Examine analyte names ...'
    #Loop through xls file
    for f in range(len(xls_list)):
        xls_name = data_dir + xls_list[f]
            
        #Open and exam each xls file
        wb = load_workbook(filename = xls_name)
        ws = wb[wb.sheetnames[0]]
            
        for c in range(20, ws.max_column + 1):
            if str(ws.cell(row = 1, column = c).value) not in element_list:
                print '    ' + xls_name + ': ' + str(ws.cell(row = 1, column = c).value) + ' analyte not in DB' 
    #=============================== 3. Verify analyte unit against database ===========================
    print '\n3. Examine unit name ...'
    unit_list = list()  #Build a list of unit name using current database values
    cur.execute("""select name from code_unit""")
    val_rows = cur.fetchall()
    for i in range(len(val_rows)):
        unit_list.append(str(val_rows[i][0]))
    unit_list.pop(0)    #Remove 1st item: 'unknown'

    #Loop through xls file
    for f in range(len(xls_list)):
        xls_name = data_dir + xls_list[f]
            
        #Open and exam each xls file
        wb = load_workbook(filename = xls_name)
        ws = wb[wb.sheetnames[0]]
            
        for c in range(20, ws.max_column + 1):
            if str(ws.cell(row = 2, column = c).value) not in unit_list:
                print '    ' + xls_name + ': ' + str(ws.cell(row = 2, column = c).value) + ' unit not in DB' 
    #==================================== 4. Check detection limit ======================================
    #Detection limit is not mandatory, which can be left blank.
    #Loop through xls files
    print '\n4. Examine detection limit ...'
    for f in range(len(xls_list)):
        xls_name = data_dir + xls_list[f]
            
        #Open and exam each xls file
        wb = load_workbook(filename = xls_name)
        ws = wb[wb.sheetnames[0]]
            
        for c in range(20, ws.max_column + 1):
            dlimit_cell = str(ws.cell(row = 3, column = c).value)
            analyte_cell = str(ws.cell(row = 1, column = c).value)
            if (dlimit_cell == 'None') or dlimit_cell.isspace():
                if analyte_cell not in nolimit_list:
                    print '    ' + xls_name + ': ' + dlimit_cell + ' missing detection limit' 
    #==================================== 5. Check method_id ======================================
    print '\n5. Examine method_id ...'
    method_list = list()   #Build a list of method_id using current database values
    cur.execute("""select method_id from code_method""")
    val_rows = cur.fetchall()
    for i in range(len(val_rows)):
        method_list.append(str(val_rows[i][0]))
    method_list.pop(0)     #Remove 1st item: 'unknown'
       
    #Loop through xls files
    for f in range(len(xls_list)):
        xls_name = data_dir + xls_list[f]
            
        #Open and exam each xls file
        wb = load_workbook(filename = xls_name)
        ws = wb[wb.sheetnames[0]]
            
        for c in range(20, ws.max_column + 1):
            if str(ws.cell(row = 4, column = c).value) not in method_list:
                print '    ' + xls_name + ': ' + str(ws.cell(row = 4, column = c).value) + ' not in DB' 
    #==================================== 6. Check lab_id ======================================
    print '\n6. Examine lab_id ...'
    lab_list = list()  #Build a list of lab using current database values
    cur.execute("""select lab_id from code_lab""")
    val_rows = cur.fetchall()
    for i in range(len(val_rows)):
        lab_list.append(str(val_rows[i][0]))
    
    #Loop through xls files
    for f in range(len(xls_list)):
        xls_name = data_dir + xls_list[f]
            
        #Open and exam each xls file
        wb = load_workbook(filename = xls_name)
        ws = wb[wb.sheetnames[0]]
            
        for c in range(20, ws.max_column + 1):
            if str(ws.cell(row = 5, column = c).value) not in lab_list:
                print '    ' + xls_name + ': ' + str(ws.cell(row = 5, column = c).value) + ' not in DB' 
    #==================================== 7. Check size_fraction ===================================
    print '\n7. Examine size_fraction ...'
    #Loop through xls files
    for f in range(len(xls_list)):
        xls_name = data_dir + xls_list[f]
            
        #Open and exam each xls file
        wb = load_workbook(filename = xls_name)
        ws = wb[wb.sheetnames[0]]
            
        for c in range(20, ws.max_column + 1):
            if (str(ws.cell(row = 6, column = c).value) == '') or (str(ws.cell(row = 6, column = c).value) == None):
                print '    ' + xls_name + ': blank size_fraction' 
    #================= 8. Check duplicate columns (i.e. same analyte, method, and lab) =============
    # It should be noted that duplicate columns may exist. For example, an analyte was initally
    # and re-analyzed using the same method by the same lab and with the same size fration. This will
    # result in 2 columns for this analyte. Extra caution should be excersized when dealing with these
    # cases.
    print '\n8. Examine duplicate comlumns within each xls file ...'    
    #Loop through xls files
    for f in range(len(xls_list)):
        xls_name = data_dir + xls_list[f]
            
        #Open and exam each xls file
        wb = load_workbook(filename = xls_name)
        ws = wb[wb.sheetnames[0]]

        work_analyte = []
        work_method = []
        work_lab = []
        work_size = []
        for c in range(20, ws.max_column + 1):
            work_analyte.append(str(ws.cell(row = 1, column = c).value))
            work_method.append(str(ws.cell(row = 4, column = c).value))
            work_lab.append(str(ws.cell(row = 5, column = c).value))
            work_size.append(str(ws.cell(row = 6, column = c).value))

        temp_analyte = list(work_analyte)
        for i in range(len(work_analyte)):
            temp_analyte[i] = str(i)
            if work_analyte[i] in temp_analyte:
                indx = temp_analyte.index(work_analyte[i])
                if (work_method[i] == work_method[indx]) and (work_lab[i] == work_lab[indx]) and \
                   (work_size[i] == work_size[indx]):
                    print '    ' + xls_name + ': ' + work_analyte[i] + ' duplicate' 
    #======================== 9. local Sample_Code problems (dupicate or blank) ========================
    print '\n9. Examine duplicate or blank Sample_Code within each xls file ...'
    #Loop through xls files
    for f in range(len(xls_list)):
        xls_name = data_dir + xls_list[f]
            
        #Open and exam each xls file
        wb = load_workbook(filename = xls_name)
        ws = wb[wb.sheetnames[0]]
            
        work_list = list()
        for r in range(8, ws.max_row + 1):
            work_cell = str(ws.cell(row = r, column = 2).value)
            if (work_cell == '') or (work_cell == ' ') or (work_cell == 'None'):
                print '    ' + xls_name + ': ' + work_cell + ' blank sample name' 
            elif  work_cell not in work_list:
                work_list.append(work_cell)
            else:
                print '    ' + xls_name + ': ' + work_cell + ' duplicate' 
    #============================ 10. Check x_coord, y_coord, z_coord, and epsg_srid ==========================
    print '\n10.Examine x-coord, y-coord, z-coord and epsg_srid ...'
    #Loop through xls files
    for f in range(len(xls_list)):
        xls_name = data_dir + xls_list[f]
            
        #Open and exam each xls file
        wb = load_workbook(filename = xls_name)
        ws = wb[wb.sheetnames[0]]

        for r in range(8, ws.max_row + 1):
            
            x_cell = str(ws.cell(row = r, column = 14).value).replace(' ', '')
            if not is_number(x_cell):
                print '    ' + xls_name + ': Row = ' + str(r) + ' ' + str(ws.cell(row = r, column = 14).value) + ' invalid x-coord'

            y_cell = str(ws.cell(row = r, column = 15).value).replace(' ', '')
            if not is_number(y_cell):
                print '    ' + xls_name + ': Row = ' + str(r) + ' ' + str(ws.cell(row = r, column = 15).value) + ' invalid y-coord'

            z_cell = str(ws.cell(row = r, column = 16).value).replace(' ', '')
            if not ((z_cell == '') or (z_cell == 'None')):
                if (not is_number(z_cell)):
                    print '    ' + xls_name + ': Row = ' + str(r) + ' ' + str(ws.cell(row = r, column = 16).value) + ' invalid z-coord'
            
            epsg_cell = str(ws.cell(row = r, column = 17).value).replace(' ', '')
            if not is_number(epsg_cell):
                print '    ' + xls_name + ': Row = ' + str(r) + ' ' + str(ws.cell(row = r, column = 17).value) + ' invalid epsg_srid' 
    #====================================== 11. Check pub_issue =======================================
    print '\n11.Examine pub_issue ...'
    #Loop through xls files
    for f in range(len(xls_list)):
        xls_name = data_dir + xls_list[f]
            
        #Open and exam each xls file
        wb = load_workbook(filename = xls_name)
        ws = wb[wb.sheetnames[0]]

        for r in range(8, ws.max_row + 1):
            work_cell = str(ws.cell(row = r, column = 18).value).replace(' ', '')
            if work_cell <> xls_list[f][:-5]:
                print '    ' + xls_name + ': ' + work_cell + ' not match file name' 
    #====================================== 12. Check coord_conf =====================================
    print '\n12.Examine Coord_Conf ...'
    #Loop through xls file under year directory
    for f in range(len(xls_list)):
        xls_name = data_dir + xls_list[f]
            
        #Open and exam each xls file
        wb = load_workbook(filename = xls_name)
        ws = wb[wb.sheetnames[0]]

        for r in range(8, ws.max_row + 1):
            work_cell = (str(ws.cell(row = r, column = 19).value).replace(' ', '')).lower()
            if (work_cell <> 'l') and (work_cell <> 'm') and (work_cell <> 'h'):
                print '    ' + xls_name + ': ' + work_cell + ' invalid coord_conf' 
    #===================================== 13. Check method-dependent unit =================================
    print '\n13.Examine method-dependent analyte unit ...'
    method_list = list()    #Build a list of method_id using current database values
    cur.execute("""select method_id from code_method""")
    val_rows = cur.fetchall()
    for i in range(len(val_rows)):
        method_list.append(str(val_rows[i][0]))
    
    group_list = list()     #Build a list of method_group using current database values
    cur.execute("""select method_group from code_method""")
    val_rows = cur.fetchall()
    for i in range(len(val_rows)):
        group_list.append(str(val_rows[i][0]))

    #Loop through xls file under year directory
    for f in range(len(xls_list)):
        xls_name = data_dir + xls_list[f]
            
        #Open and exam each xls file
        wb = load_workbook(filename = xls_name)
        ws = wb[wb.sheetnames[0]]
            
        for c in range(20, ws.max_column + 1):
            analyte_cell = str(ws.cell(row = 1, column = c).value).replace(' ', '')
            unit_cell = str(ws.cell(row = 2, column = c).value).replace(' ', '')
            method_cell = str(ws.cell(row = 4, column = c).value).replace(' ', '')

            right_unit = standard_unit[element_list.index(analyte_cell)]
            check_code = group_list[method_list.index(method_cell)]
            if analyte_cell in dependent_analyte:
                if check_code == dependent_method[dependent_analyte.index(analyte_cell)]:
                    right_unit = dependent_unit[dependent_analyte.index(analyte_cell)]
                        
            if unit_cell <> right_unit:
                print '    ' + xls_name + ': ' + analyte_cell + ' ' + unit_cell + ' should be ' + right_unit 
    #======================================== 14. Check fixed analyte-method combo ====================================
    print '\n14.Examine fixed analyte-method combo ...'
    #Loop through xls file under year directory
    for f in range(len(xls_list)):
        xls_name = data_dir + xls_list[f]
            
        #Open and exam each xls file
        wb = load_workbook(filename = xls_name)
        ws = wb[wb.sheetnames[0]]
            
        for c in range(20, ws.max_column + 1):
            analyte_cell = str(ws.cell(row = 1, column = c).value).replace(' ', '')
            method_cell = str(ws.cell(row = 4, column = c).value).replace(' ', '')
            if analyte_cell in fix_analyte:
                right_method = fix_method[fix_analyte.index(analyte_cell)]
                if method_cell <> right_method:
                    print '    ' + xls_name + ': ' + analyte_cell + ' ' + method_cell + ' should be ' + right_method 
    #========================================== 15. Check analyte value =================================================
    print '\n15.Examine analyte values ...'
    #Loop through xls file under year directory
    for f in range(len(xls_list)):
        xls_name = data_dir + xls_list[f]
            
        #Open and exam each xls file
        wb = load_workbook(filename = xls_name)
        ws = wb[wb.sheetnames[0]]
            
        for c in range(20, ws.max_column + 1):
            analyte_cell = str(ws.cell(row = 1, column = c).value).replace(' ', '')
            unit_cell = str(ws.cell(row = 2, column = c).value).replace(' ', '')

            for r in range(8, ws.max_row + 1):
                analyte_value = str(ws.cell(row = r, column = c).value)
                
                if analyte_value == 'None':
                    continue
                elif analyte_value == '0':
                    print '    ' + xls_name + '->' + analyte_cell + ': ' + \
                            analyte_value + '(' + str(r) + ',' + str(c) + ') wrong analyte value'
                elif (not is_number(analyte_value)):
                    if (analyte_value[0] <> '>') and (analyte_value[0] <> '<'):
                        print '    ' + xls_name + '->' + analyte_cell + ': ' + \
                        analyte_value + '(' + str(r) + ',' + str(c) + ') wrong analyte value'
                elif (unit_cell == '%') and (is_number(analyte_value)):
                    #Greater-than-100% value is not allowed if anayte unit is % 
                    if (analyte_cell <> 'Total') and float(analyte_value) > 100.0:
                        print '    ' + xls_name + '->' + analyte_cell + ': ' + analyte_value + ' > 100%'

        #Examine rows without any analytic values
        for r in range(8, ws.max_row + 1):
            val_count = 0
            for c in range(20, ws.max_column + 1):
                if is_number(str(ws.cell(row = r, column = c).value).replace(' ', '')):
                    val_count = val_count + 1
            if val_count == 0:
                print '    ' + xls_name + ': Row = ' + str(r) + ' Blank row with any analytic values' 
    #================== 16. Check global duplicate samples among the xlsx files being screened =============
    '''print '\n16.Examine duplicate samples among those in the xlsx files  ...' 
    #Looking for samples which are closely located; having similar sample names, and in different publications
    #Loop through the xls files
    xls_sample = list()  #Build a list for sample_code in xlsx files
    xls_xc = list()      #Build a list for x_coord in xlsx files
    xls_yc = list()      #Build a list for y_coord in xlsx files
    xls_pub = list()     #Build a list for pub_issue in xlsx files
    xls_file = list()    #Build a list for xlsx file names
    for f in range(len(xls_list)):
        xls_name = data_dir + xls_list[f]
            
        #Open and exam each xls file
        wb = load_workbook(filename = xls_name)
        ws = wb[wb.sheetnames[0]]

        for r in range(8, ws.max_row + 1):
            cell_sample = str(ws.cell(row = r, column = 2).value).replace(' ', '')
            cell_xc = str(ws.cell(row = r, column = 14).value).replace(' ', '')
            cell_yc = str(ws.cell(row = r, column = 15).value).replace(' ', '')
            cell_epsg = str(ws.cell(row = r, column = 17).value).replace(' ', '')
            cell_pub = str(ws.cell(row = r, column = 18).value).replace(' ', '')
            
            xls_sample.append(cell_sample)
            
            if cell_epsg <> '4269': #NAD83 geographic
                [cell_xc, cell_yc] = project2nad83 (float(cell_xc), float(cell_yc), int(cell_epsg))
            xls_xc.append(cell_xc)
            xls_yc.append(cell_yc)
            xls_pub.append(cell_pub)
            xls_file.append(xls_name)
    
    #Loop through the collected samples
    for i in range(len(xls_sample) - 1):
        for j in range(i + 1, len(xls_sample)):
            #Examine sample location closeness
            pos_match = 0
            if (abs(float(xls_xc[j]) - float(xls_xc[i])) <= min_lon) and \
               (abs(float(xls_yc[j]) - float(xls_yc[i])) <= min_lat):
                pos_match = 1

            #Examine sample name closeness
            short_sample = xls_sample[j]
            long_sample = xls_sample[i]
            if len(xls_sample[j]) > len(xls_sample[i]):
                short_sample = xls_sample[i]
                long_sample = xls_sample[j]
            c_counter = 0
            for c in range(len(short_sample)):
                if short_sample[c] in long_sample:
                    c_counter = c_counter + 1
               
                    #Remove the matched character from the "long_sample"
                    c_indx = long_sample.index(short_sample[c]) 
                    long_sample = long_sample[0:c_indx] + long_sample[(c_indx + 1):]
                    
            name_match = 0
            if c_counter/float(len(short_sample)) > 0.9:
                name_match = 1
            
            if (pos_match == 1) and (name_match == 1):
                print '    Duplicate samples: ' + xls_file[j] + '-' + xls_sample[j] + ' -> ' + \
                                                  xls_file[i] + '-' + xls_sample[i] '''
    #=============== 17. Check global duplicate samples between database and the xlsx files =================
    '''print '\n17.Examine duplicate samples between those in database and in the xlsx files  ...' 
    #Add sample_code, x_coord, y_coord, epsg_snrd, and pub_issue (to the lists) in the xlsx files being screened 
    #Collect data in the xls files
    xls_sample = list()  #Build a list for sample_code in xlsx files
    xls_xc = list()      #Build a list for x_coord in xlsx files
    xls_yc = list()      #Build a list for y_coord in xlsx files
    xls_pub = list()     #Build a list for pub_issue in xlsx files
    xls_file = list()    #Build a list for xlsx file names
    
    for f in range(len(xls_list)):
        xls_name = data_dir + xls_list[f]
            
        #Open and exam each xls file
        wb = load_workbook(filename = xls_name)
        ws = wb[wb.sheetnames[0]]

        for r in range(8, ws.max_row + 1):
            cell_sample = str(ws.cell(row = r, column = 2).value).replace(' ', '')
            cell_xc = str(ws.cell(row = r, column = 14).value).replace(' ', '')
            cell_yc = str(ws.cell(row = r, column = 15).value).replace(' ', '')
            cell_epsg = str(ws.cell(row = r, column = 17).value).replace(' ', '')
            cell_pub = str(ws.cell(row = r, column = 18).value).replace(' ', '')
            
            xls_sample.append(cell_sample)
            
            if cell_epsg <> '4269': #NAD83 geographic
                [cell_xc, cell_yc] = project2nad83 (float(cell_xc), float(cell_yc), int(cell_epsg))
            xls_xc.append(cell_xc)
            xls_yc.append(cell_yc)
            xls_pub.append(cell_pub)
            xls_file.append(xls_name)
            
    #Collect data in database
    db_sample = list()  #Build a list for sample_code in database
    db_xc = list()      #Build a list for x_coord in database
    db_yc = list()      #Build a list for y_coord in database
    db_epsg = list()    #Build a list for EPSG_SRID in database

    cur.execute("""select sample_code, x_coord, y_coord, EPSG_SRID from data_sample""")
    val_rows = cur.fetchall()
    for i in range(len(val_rows)):
        db_sample.append(val_rows[i][0])
        
        if str(val_rows[i][3]) <> '4269': #NAD83 geographic
            [cell_xc, cell_yc] = project2nad83 (float(val_rows[i][1]), float(val_rows[i][2]), int(val_rows[i][3]))
            db_xc.append(cell_xc)
            db_yc.append(cell_yc)
        else:
            db_xc.append(float(val_rows[i][1]))
            db_yc.append(float(val_rows[i][2]))
        
    #Compare each sample in the xls files with those in the database
    for i in range(len(xls_sample)):
        for j in range(len(db_sample)):
            #Examine sample location closeness
            pos_match = 0
            if (abs(float(db_xc[j]) - float(xls_xc[i])) <= min_lon) and \
               (abs(float(db_yc[j]) - float(xls_yc[i])) <= min_lat):
                pos_match = 1

            #Examine sample name closeness
            short_sample = db_sample[j]
            long_sample = xls_sample[i]
            if len(db_sample[j]) > len(xls_sample[i]):
                short_sample = xls_sample[i]
                long_sample = db_sample[j]
            c_counter = 0
            for c in range(len(short_sample)):
                if short_sample[c] in long_sample:
                    c_counter = c_counter + 1
               
                    #Remove the matched character from the "long_sample"
                    c_indx = long_sample.index(short_sample[c]) 
                    long_sample = long_sample[0:c_indx] + long_sample[(c_indx + 1):]
                    
            name_match = 0
            if c_counter/float(len(short_sample)) > 0.9:
                name_match = 1
                        
            if (pos_match == 1) and (name_match == 1):
                print '    Duplicate samples: ' + 'DB - ' + db_sample[j] + ' -> ' + xls_file[i] + '-' + xls_sample[i] '''
    
    db_conn.close()
    print '\nJob done.' 
    
if __name__ == "__main__":
    main()
