# -*- coding: cp1252 -*-
'''+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
   This script is modified from Tian Han's tillDB_data_screener script. It is coded to identify and report problems
   found in the staged ARIS geochem data certificates xlsx files   before loading them loaded into the database.
   Please beware that this script only identifies and reports problems. The problems identified need to be fixed
   manually.
   
   Input
         1) Path to the directory where the staged certificate xlsx files reside
         2) ARIS Geochem stage Database path
         3) Some hardcoded parameters
         
   Output
         Display identified problems

   Operation note
         1) The script has many independent code-blocks targeting different types of errors as listed
            below. Though each code-block can be executed randomly, it is strongly recommended to
            run them sequentially in the given order, one at a time. For example, it should start with
            checking format related problems in the xlsx files by commenting out all other code-blocks
            which are for checking other types of errors. Once the format-related errors are identified
            and fixed, the format-checking code-block should be commented out and then move on to the
            "analyte name checking block" ...

         2) Duplicate samples should be given different "sample_code". Re-published samples should
            have the same 'sample_code' across all publications. Sections 15 and 16 are designed to
            ensure if the above requirements are met.
            
    This script is able to check the following:

    
  Status
      In Development

  Furture improment


  Developer
      Gabe Fortin
      
  Last update
      2017-06-08
  +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++'''
import os, sys, csv, pyodbc, ogr, osr
from openpyxl import load_workbook
from dateutil.parser import parse

# This function is to check if a given string can be converted to a decimal number
# Syntax: is_number(string) return logic
def is_number(n):
    try:
        float(n)
        return True
    except ValueError:
        return False
    
# This function is to check if a given variable can be converted to a string
# Syntax: is_number(variable) return logic
def is_string(s):
    try:
        str(s)
        return True
    except ValueError:
        return False
    
#Project input coordinates to NAD83 lat. and long (EPSG_SRID = 4269)
#Syntax: project2nad83(float, float, int) returns [nad83_long, nad83_lat]
def project2nad83 (x_coord, y_coord, source_epsg):

    #Create a point geometry using the given coordinates
    point = ogr.Geometry(ogr.wkbPoint)
    point.AddPoint(x_coord, y_coord)

    #Specify source spatial reference and projection
    source_ref = osr.SpatialReference()
    source_ref.ImportFromEPSG(source_epsg)

    #Define target spatial reference and projection based on input
    target_ref = osr.SpatialReference()
    target_ref.ImportFromEPSG(4269)
        
    #Perform transformation
    transform = osr.CoordinateTransformation(source_ref, target_ref)
    point.Transform(transform)
    
    return [str(point.GetX(0)), str(point.GetY(0))]
    
def main():   
    #File path
    db_path = 'C:\\Project\\ARIS_Geochem_dev\\data\\ARIS_geochem_stage.accdb'
    data_dir = 'C:\\Project\\ARIS_Geochem_dev\\data\\_AR Data Staging Certificate\\'

    #Database connection
    db_conn = pyodbc.connect('Driver={Microsoft Access Driver (*.mdb, *.accdb)};DBQ='+db_path)
    cur = db_conn.cursor()

    #Collect year sub-directories under given "data_dir"
    xls_list = os.listdir(data_dir)
    #====================================== 1. Check xls file format ======================================

    print '1. Examine file format ...'
    #Loop through xls file
    for f in range(len(xls_list)):
        xls_name = data_dir + xls_list[f]
            
        #Open and exam each xls file
        wb = load_workbook(filename = xls_name)
        ws = wb[wb.sheetnames[0]]
        if (str(ws.cell(row = 1, column = 1).value)).lower() <> 'cert_no':
            print '    ' + xls_list[f] + ': Wrong cert_no column header'
        if (str(ws.cell(row = 1, column = 2).value)).lower() <> 'cert_date':
            print '    ' + xls_list[f] + ': Wrong cert_date column header'
        if (str(ws.cell(row = 1, column = 3).value)).lower() <> 'lab_id':
            print '    ' + xls_list[f] + ': Wrong lab_id column header'
        if (str(ws.cell(row = 1, column = 4).value)).lower() <> 'prep_id':
            print '    ' + xls_list[f] + ': Wrong pre_id column header'

    #=============================== 2. Verify lab_id ===========================
    print '\n2. Examine lab_id...'
    lab_list = list()  #Build a list of lab id's using current database values
    cur.execute("""select lab_id from code_lab""")
    val_rows = cur.fetchall()
    for i in range(len(val_rows)):
        lab_list.append(str(val_rows[i][0]))

    #Loop through xls file
    for f in range(len(xls_list)):
        xls_name = data_dir + xls_list[f]
            
        #Open and exam each xls file
        wb = load_workbook(filename = xls_name)
        ws = wb[wb.sheetnames[0]]

        if str(ws.cell(row = 2, column = 3).value) not in lab_list:
            print '    ' + xls_list[f] + ': ' + str(ws.cell(row = 2, column = 3).value) + \
                    ' lab_id not in list'

    #==================================== 3. Check prep_id ======================================
    print '\n3. Examine prep_id...'
    prep_list = list()  # Build a list of lab id's using current database values
    cur.execute("""select prep_id from code_prep""")
    val_rows = cur.fetchall()
    for i in range(len(val_rows)):
        prep_list.append(str(val_rows[i][0]))

    # Loop through xls file
    for f in range(len(xls_list)):
        xls_name = data_dir + xls_list[f]

        # Open and exam each xls file
        wb = load_workbook(filename=xls_name)
        ws = wb[wb.sheetnames[0]]

        if str(ws.cell(row=2, column=4).value) not in prep_list:
            print '    ' + xls_list[f] + ': ' + str(ws.cell(row=2, column=4).value) + \
                  ' prep_id not in list'

    # ----------------------------------4.Check date validity-----------------------------------------------------
    print '\n4.Examine sample dates ...'
    # Loop through xls files
    for f in range(len(xls_list)):
        xls_name = data_dir + xls_list[f]

        # Open and exam each xls file
        wb = load_workbook(filename=xls_name)
        ws = wb[wb.sheetnames[0]]

        date_cell = str(ws.cell(row=2, column=2).value)

        if not ((date_cell == '') or (date_cell == 'None')):
            if not (date_cell.isupper() or (date_cell.islower())):
                try:
                    parse(date_cell)
                except ValueError:
                    print '    ' + xls_name + ': ' + str(ws.cell(row=2, column=2).value) + \
                            ' is not a proper date'
            else:
                print '    ' + xls_name + ': ' + str(ws.cell(row=2, column=2).value) + \
                        ' is not a proper date'
        else:
            print '    ' + xls_name + ': ' + str(ws.cell(row=2, column=2).value) + \
                  ' certificate date is missing'

    # =============================== 5. Verify cert_no ===========================
    print '\n5. Examine cert_no...'
    cert_list = list()  # Build a list of lab id's using current database values
    cur.execute("""select cert_no from data_cert""")
    val_rows = cur.fetchall()
    for i in range(len(val_rows)):
        cert_list.append(str(val_rows[i][0]))

    # Loop through xls file
    for f in range(len(xls_list)):
        xls_name = data_dir + xls_list[f]

        # Open and exam each xls file
        wb = load_workbook(filename=xls_name)
        ws = wb[wb.sheetnames[0]]

        if str(ws.cell(row=2, column=1).value) in cert_list:
            print '    ' + xls_list[f] + ': ' + str(ws.cell(row=2, column=1).value) + \
                    ' cert_no already in db'
        elif str(ws.cell(row=2, column=1).value) == 'None':
            print '    ' + xls_list[f] + ': ' + str(ws.cell(row=2, column=1).value) + \
                    ' cert_no missing'

        # extract the cert_no from the file name and compare to that in the sheet
        cert_no = xls_list[f].partition('_')[2]
        cert_no = cert_no[:-10]
        if cert_no <> str(ws.cell(row=2, column=1).value):
            print '    ' + xls_list[f] + ': ' + str(ws.cell(row=2, column=1).value) + \
                    ' cert_no in file name does not match cert_no within sheet' + cert_no


'''   
db_conn.close()
'''
print '\nJob done.'

if __name__ == "__main__":
    main()
